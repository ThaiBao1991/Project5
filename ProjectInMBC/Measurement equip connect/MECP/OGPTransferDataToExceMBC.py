# Đây là phiên bản đã có khả năng lướt qua file csv, tự động load dữ liệu của form khá ổn.
# Có các tính năng cơ bản sau:
# - Khả năng tự động load chọn form mà có dữ liệu data_config khi click vào nút update. Nút update có khả năng update folder, file chạy, form cần xuất tương ứng với form
# - Khả năng hỏi khi save và đã thêm dateTime để click ok là lưu file
# - Khả năng Update code nhưng chưa sử dụng tới, có khả năng nếu trong form có lưu code thì sẽ load file thực thi chạy code 
# - Nhấn nút close là thoát
# - Nhấp phím tắt ALt + S sẽ hiện ra bảng login, nếu login thành công sẽ vào bảng admin, dù đang trống nhưng có khả năng điều chỉnh trong tương lai

#Khai báo thư viện:
from datetime import datetime # Lib ngày giờ hệ thống
import os       # Lib xử lý tập tin, thư mục ....liên quan hệ điều hành
import openpyxl  # Lib xử lý file excel
import PySimpleGUI as sg    # Lib giao diện
import re #regex
import json # dùng để tạo các file json và đọc
import textwrap # dùng để chỉnh các text
import shutil # dùng để copy dữ liệu có kèm metadata có thể sử dụng thư viên os cũng ok mà không có metadata
import pandas as pd
import difflib #Check 2 File String
import sys #Kiểm tra tình trạng chạy file
# Basic function
# Kiểm tra xem chương trình đang chạy dưới dạng mã nguồn Python hay là file .exe
folder_save = None
if getattr(sys, 'frozen', False):
    # Chương trình đang chạy dưới dạng file .exe

    # Lấy đường dẫn tới thư mục hiện hành
    current_directory = sys._MEIPASS

    # Tạo đường dẫn tới file data_form_config.txt trong môi trường chạy của chương trình
    source_file_path_pyinstaller = os.path.join(current_directory, 'data_form_config.txt')

    # Tạo đường dẫn tới file data_form_config.txt trong thư mục hiện tại
    destination_file_path_pyinstaller = os.path.join(os.getcwd(), 'data_form_config.txt')

    # Kiểm tra xem file đã tồn tại chưa
    if not os.path.exists(destination_file_path_pyinstaller):
        # Nếu file chưa tồn tại, sao chép file
        shutil.copy(source_file_path_pyinstaller, destination_file_path_pyinstaller)

# Đọc dữ liệu từ file
def read_data():
    try:
        with open('data_form_config.txt', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

# Lưu dữ liệu vào file
def save_data(data):
    with open('data_form_config.txt', 'w') as f:
        json.dump(data, f)

def copy_from_excel_to_excel_horizontal(source_path, range1, destination_path, range2, number_of_copy):
    # print("start copy excel to excel horizontal")
    # Load the source workbook and get the active sheet
    source_workbook = openpyxl.load_workbook(source_path)
    source_sheet = source_workbook.active
    
    # Load the destination workbook and get the active sheet
    destination_workbook = openpyxl.load_workbook(destination_path)
    destination_sheet = destination_workbook.active
        
    # Calculate the source range for copying
    resultSource = get_col_row(range1)
    src_col_start = resultSource[0]
    src_row_start = int(resultSource[1])
    src_range = source_sheet[src_col_start + str(src_row_start):chr(ord(src_col_start) + number_of_copy - 1) + str(src_row_start)]
    
    # print("src_range =" ,src_range)
    # Calculate the starting position for pasting
    resultDestination = get_col_row(range2)
    dst_col_start = resultDestination[0]
    dst_row_start = int(resultDestination[1])

    # Copy each cell from the source range to the destination range
    for col in src_range:
        for cell_index, cell in enumerate(col, start=0):
            dst_cell = destination_sheet[chr(ord(dst_col_start) + cell_index) + str(dst_row_start)]
            dst_cell.value = cell.value

    # Save the destination workbook
    destination_workbook.save(destination_path)
    # print("End copy excel to excel horizontal")
    
def copy_from_excel_to_excel_vertical(source_path, source_range, destination_path, destination_range, number_of_copy):
    # Load the source workbook
    source_wb = openpyxl.load_workbook(source_path)
    source_ws = source_wb.active

    # Load the destination workbook
    destination_wb = openpyxl.load_workbook(destination_path)
    destination_ws = destination_wb.active

    # Determine the column and the starting row of the source range
    resultSource = get_col_row(source_range)
    src_col = resultSource[0]
    src_row_start = int(resultSource[1])

    # Copy the cells from the source to the destination
    for i in range(number_of_copy):
        # Get the source cell value
        src_cell = source_ws[src_col + str(src_row_start + i)].value
        # Determine the destination cell reference
        resultDestination = get_col_row(destination_range)
        dst_col = resultDestination[0]
        dst_row_start = int(resultDestination[1])
        # Set the value to the destination cell
        destination_ws[dst_col + str(dst_row_start + i)].value = src_cell

    # Save the destination workbook
    destination_wb.save(destination_path)

def copy_horizontal_to_vertical(source_path, range1, destination_path, range2, number_of_copy):
    # Load the source workbook
    source_wb = openpyxl.load_workbook(source_path)
    source_ws = source_wb.active
 
    # Load the destination workbook
    destination_wb = openpyxl.load_workbook(destination_path)
    destination_ws = destination_wb.active
 
    # Determine the starting point in the source sheet
    resultSource = get_col_row(range1)
    src_col_start = resultSource[0]
    src_row_start = int(resultSource[1])
 
    # Determine the starting point in the destination sheet
    resultDestination = get_col_row(range2)
    dest_col_start = resultDestination[0]
    dest_row_start = int(resultDestination[1])
 
    # Perform the copy from horizontal to vertical
    for i in range(number_of_copy):
        # Get the value from the source cell
        src_cell_value = source_ws[chr(ord(src_col_start) + i) + str(src_row_start)].value
        # Set the value in the destination cell
        destination_ws[dest_col_start + str(dest_row_start + i)].value = src_cell_value
 
    # Save the changes to the destination workbook
    destination_wb.save(destination_path)

def copy_vertical_to_horizontal(source_path, range1, destination_path, range2, number_of_copy):
    # Load the source workbook
    source_wb = openpyxl.load_workbook(source_path)
    source_ws = source_wb.active
 
    # Load the destination workbook
    destination_wb = openpyxl.load_workbook(destination_path)
    destination_ws = destination_wb.active
 
    # Determine the starting point in the source sheet
    resultSource = get_col_row(range1)
    src_col = resultSource[0]
    src_row_start = int(resultSource[1])
 
    # Determine the starting point in the destination sheet
    resultDestination = get_col_row(range2)
    dest_col_start = resultDestination[0]
    dest_row_start = int(resultDestination[1])
 
    # Perform the copy from vertical to horizontal
    for i in range(number_of_copy):
        # Get the value from the source cell
        src_cell_value = source_ws[src_col + str(src_row_start + i)].value
        # Set the value in the destination cell
        destination_ws[chr(ord(dest_col_start) + i) + str(dest_row_start)].value = src_cell_value
 
    # Save the changes to the destination workbook
    destination_wb.save(destination_path)
    
def get_col_row(cell_range):
    match = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", cell_range)
    if match:
        return match.groups()
    else:
        match = re.match(r"([A-Za-z]+)(\d+)", cell_range)
        if match:
            return match.group(1), match.group(2), match.group(1), match.group(2)
            
        else:
            return None, None, None, None

def copy_excel_transpose(source_path, range1, destination_path, range2):
    # Load the source workbook
    source_wb = openpyxl.load_workbook(source_path)
    source_ws = source_wb.active

    # Load the destination workbook
    destination_wb = openpyxl.load_workbook(destination_path)
    destination_ws = destination_wb.active

    # Determine the starting and ending points in the source sheet
    src_col_start, src_row_start, src_col_end, src_row_end = get_col_row(range1)

    # Determine the starting point in the destination sheet
    resultDestination = get_col_row(range2)
    dest_col_start = resultDestination[0]
    dest_row_start = int(resultDestination[1])

    # Check if the source data is horizontal, vertical, or a single cell
    if ord(src_col_end) - ord(src_col_start) > int(src_row_end) - int(src_row_start):  # Horizontal
        # Perform the copy from horizontal to vertical
        for j in range(int(src_row_end) - int(src_row_start) + 1):
            for i in range(ord(src_col_end) - ord(src_col_start) + 1):
                # Get the value from the source cell
                src_cell_value = source_ws[chr(ord(src_col_start) + i) + str(int(src_row_start) + j)].value
                # Set the value in the destination cell
                destination_ws[dest_col_start + str(dest_row_start+i)].value = src_cell_value
            # Update the starting point in the destination sheet for the next row
            dest_col_start = chr(ord(dest_col_start) + 1)
    elif ord(src_col_end) - ord(src_col_start) < int(src_row_end) - int(src_row_start):  # Vertical
        # print("Vertical to horizontal")
        # Perform the copy from vertical to horizontal
        for i in range(ord(src_col_end) - ord(src_col_start) + 1):
            for j in range(int(src_row_end) - int(src_row_start) + 1):
                print(i,j,chr(ord(dest_col_start) + i + j * (int(src_row_end) - int(src_row_start) + 1)) + str(dest_row_start))
                # Get the value from the source cell
                src_cell_value = source_ws[chr(ord(src_col_start) + i) + str(int(src_row_start) + j)].value
                # Set the value in the destination cell
                destination_ws[chr(ord(dest_col_start) + j) + str(dest_row_start)].value = src_cell_value
            # Update the starting point in the destination sheet for the next column
            dest_row_start += 1
    else:  # Single cell
        print("Copy single cell")
        # Copy the value directly
        src_cell_value = source_ws[src_col_start + str(src_row_start)].value
        destination_ws[dest_col_start + str(dest_row_start)].value = src_cell_value

    # Save the changes to the destination workbook
    destination_wb.save(destination_path)
def main():
    sg.theme_text_color('black')
    options = ['Type1LAn2', 'Type2:an2', 'Type3', 'Type4', 'Type5', 'Type6', 'Type7', 'Type8', 'Type9', 'Type10']
    data_copy_options_layout=[[sg.Checkbox("All-Lan2",key="All2",enable_events=True)]]
    data = read_data()  # Đọc dữ liệu
    categories = list(data.keys())  # Trích xuất các Category từ dữ liệu
    for option in options:
        data_copy_options_layout.append([sg.Checkbox(option,key=option,enable_events=True)])
    Layout = [    
            [sg.Text("Chọn chương trình OGP :",font=('Helvetica', 11,'bold')),sg.Combo(values=categories,key="FormCategory",enable_events=True,size=(70,10))],
            
            [sg.Text("Chọn file dữ liệu chương trình OGP :",font=('Helvetica', 11,'bold')),sg.InputText(disabled=True,size=(50, 1), key='form_OJB'),sg.FilesBrowse(key='form_OJB_choose')],
            
            [sg.Text("Chọn file xuất dữ liệu xuất hàng ",font=('Helvetica', 11,'bold')),sg.InputText(justification='center',disabled=True,size=(50, 1), key='form_save',enable_events=True),sg.FilesBrowse(key='form_save_choose')],
            
            [sg.Text("Nhập số lượng dòng dữ liệu :",font=('Helvetica', 11,'bold')),sg.InputText(default_text="10",justification="center",disabled=False,size=(7,1),key="number_of_data")],

            # [sg.Frame('Dữ liệu cần copy',[
            #     [sg.Checkbox('All', key='All', enable_events=True)],
            #     [sg.Checkbox('Type1', key='Type1', enable_events=True)],
            #     [sg.Checkbox('Type2', key='Type2', enable_events=True)],
            #  ],key='frame1',visible=False)  
            # ],
            
            # [[sg.Frame('Dữ liệu copy 2',[data_copy_options_layout],key='frame2',visible=False)]],

            [sg.Button('Update Form',key="Start", size=(10, 2), button_color=('green', 'grey'), font=('Helvetica', 10,'bold'), pad=(10, 10),enable_events=True),
                sg.Button('Open Form',key="Open", size=(10,2),button_color=('yellow','grey'),enable_events=True,font=('Helvetica',10,'bold'),pad=(10, 10))],

            [sg.Column([[sg.Text("VDM-Inspection Section 2024/04", font=('Helvetica',8))]], justification='left',expand_x=True),
            sg.Column([[sg.Button('Close', size=(10, 2), button_color=('red', 'grey'), font=('Helvetica', 8,'bold'), pad=(0, 0), enable_events=True)]], justification='right')] 
        ]

    def create_login_window():
        layout = [[sg.Text('Username'), sg.Input(key='username')],
                [sg.Text('Password'), sg.Input(key='password', password_char='*')],
                [sg.Button('Login',key="LoginBtn",enable_events=True), sg.Button('Thoát',key="CloseLoginBtn",enable_events=True)]]
        return sg.Window('Đăng nhập', layout, return_keyboard_events=True)

    # Tạo cửa sổ admin
    def create_admin_window():
        # Sử dụng biến toàn cục
        global folder_save
        data = read_data()  # Đọc dữ liệu
        categories = list(data.keys())  # Trích xuất các Category từ dữ liệu
        layout = [
                [sg.Text("Chọn folder lưu file: ",font=('Helvetica', 11,'bold'))],
                [sg.InputText(folder_save, disabled=True, size=(50, 1), key='location_save'),
                sg.FolderBrowse(key='folder_save_choose'),
                sg.Button("Update")],
                
                [sg.Text("Chọn file dữ liệu chương trình OGP ",font=('Helvetica', 11,'bold'))],
                    [sg.InputText(disabled=True,size=(50, 1), key='form_OJB'),sg.FilesBrowse(key='form_OJB_choose'),sg.Button("Clear link",key="ClearBtn_OGP_file_link",enable_events=True)],
                    
                    [sg.Text("Chọn chương trình OGP",font=('Helvetica', 11,'bold')),sg.Combo(values=categories,key="FormCategory",enable_events=True,readonly=False,size=(35,4))],
                    
                    [sg.Text("Chọn file xuất dữ liệu xuất hàng ",font=('Helvetica', 11,'bold'))],
                    [sg.InputText(justification='center',disabled=True,size=(50, 10), key='form_save',enable_events=True),sg.FilesBrowse(key='form_save_choose'),sg.Button("Clear link",key="ClearBtn_Data_export_link",enable_events=True)],
                    
                    [sg.Text("Nhập số lượng dòng dữ liệu :",font=('Helvetica', 11,'bold')),sg.InputText(justification="center",disabled=False,size=(20,1),key="number_of_data")]
                ]
                    
        return sg.Window('Admin', layout)


    # window=sg.Window('Chương trình lấy dữ liệu từ máy OGP vào biểu ghi chép',Layout,size=(650,450),resizable=False,finalize=True)
    windowOGP=sg.Window('Chương trình lấy dữ liệu từ máy OGP vào biểu ghi chép',Layout,resizable=False,finalize=True,return_keyboard_events=True)
    windowOGP.bind('<Alt-s>', 'alt_s')

    while True:
        # event, values = window.read(timeout=20)
        window, event, values = sg.read_all_windows()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'alt_s':  # Sự kiện nhấn tổ hợp phím ALT + S
            print("Chạy ALT + S")
            login_window = create_login_window()
            while True:
                event, values = login_window.read()
                if event in (sg.WINDOW_CLOSED, 'Thoát', 'CloseLoginBtn'):
                    login_window.close()
                    break
                elif event == 'LoginBtn':
                    if values['username'] == 'bao1991' and values['password'] == 'ktbao1991':
                        sg.popup('Đăng nhập thành công!')
                        login_window.close()
                        admin_window = create_admin_window()
                        while True:  # Thêm một vòng lặp sự kiện riêng biệt cho layout admin
                            event, values = admin_window.read()
                            if event in (sg.WINDOW_CLOSED, 'Thoát', 'CloseAdminBtn'):
                                admin_window.close()
                                break
                            elif event == 'Update':
                                data = read_data()
                                categories = list(data.keys())  # Trích xuất các Category từ dữ liệu
                                new_folder_save = values['folder_save_choose']
                                new_form_OJB = values['form_OJB']
                                new_form_save = values['form_save']
                                number_of_data = values['number_of_data']
                                code = ""
                                # code = """
                                #             # print("start button click")
                                #             # copy_from_excel_to_excel_horizontal("ver1.xlsx","B7","convert ver1.xlsx","B4",10)
                                #             # copy_from_excel_to_excel_vertical("ver1.xlsx","B7","convert ver1.xlsx","B5",10)
                                #             # copy_horizontal_to_vertical("ver1.xlsx","B7","convert ver1.xlsx","C5",10)
                                #             # copy_vertical_to_horizontal("ver1.xlsx","B7","convert ver1.xlsx","C5",10)
                                #             # copy_excel_transpose("ver1.xlsx","B7:F8","convert ver1.xlsx","B4")
                                #             # copy_excel_transpose("ver1.xlsx","B7:F7","convert ver1.xlsx","B4")
                                #             copy_excel_transpose("ver1.xlsx","B7:C11","convert ver1.xlsx","B4")
                                #             copy_excel_transpose("ver1.xlsx","B7:B11","convert ver1.xlsx","B4")
                                #             copy_excel_transpose("ver1.xlsx","C7","convert ver1.xlsx","B4")
                                #         """
                                if new_folder_save:  # Người dùng đã chọn một thư mục mới
                                    data[values['FormCategory']] = {'folder_save': new_folder_save, 'form_OJB': new_form_OJB, 'form_save': new_form_save,'number_of_data':number_of_data, 'code': code}
                                else:  # Người dùng không chọn thư mục mới, giữ nguyên đường dẫn thư mục hiện tại
                                    if values['FormCategory'] not in data:
                                        data[values['FormCategory']] = {}
                                    if isinstance(data.get(values['FormCategory']), dict):  # Kiểm tra xem data[values['FormCategory']] có phải là một từ điển hay không
                                        data[values['FormCategory']]['folder_save'] = values['location_save']
                                        if new_form_OJB is not None:  # Người dùng đã chọn một file OJB mới
                                            data[values['FormCategory']]['form_OJB'] = new_form_OJB
                                        if new_form_save is not None:  # Người dùng đã chọn một file save mới
                                            data[values['FormCategory']]['form_save'] = new_form_save
                                        data[values['FormCategory']]['code'] = code  # Thêm đoạn mã vào dữ liệu
                                        if number_of_data is not None :
                                            data[values['FormCategory']]['number_of_data']= number_of_data
                                    else:
                                        sg.Popup('Vui lòng chọn một thư mục mới', keep_on_top=True)
                                save_data(data)
                                sg.popup(f"Hoàn thành update đường dẫn của  {values['FormCategory']}" ,title='Hoàn Thành')
                                categories = list(data.keys())  # Trích xuất các Category từ dữ liệu
                                # Cập nhật danh sách 'categories' trong 'sg.Combo' sau khi lưu dữ liệu
                                admin_window['FormCategory'].update(values=categories)
                            elif event == 'ClearBtn_Data_export_link':
                                    admin_window['form_save'].update('')
                            elif event == 'ClearBtn_OGP_file_link':
                                admin_window['form_OJB'].update('')
                            elif event == 'FormCategory':
                                data = read_data()
                                form_category = values['FormCategory']
                                if form_category in data:
                                    if isinstance(data[form_category], dict):  # Kiểm tra xem data[form_category] có phải là một từ điển hay không
                                        admin_window['location_save'].update(data[form_category].get('folder_save', ''))
                                        admin_window['form_OJB'].update(data[form_category].get('form_OJB', ''))
                                        admin_window['form_save'].update(data[form_category].get('form_save', ''))
                                        admin_window['number_of_data'].update(data[form_category].get('number_of_data',''))
                                    else:
                                        admin_window['location_save'].update(data[form_category])
                                        admin_window['form_OJB'].update('')
                                        admin_window['form_save'].update('')
                                        admin_window['number_of_data'].update('10')
                                else:
                                    admin_window['location_save'].update('')
                                    admin_window['form_OJB'].update('')
                                    admin_window['form_save'].update('')
                                    admin_window['number_of_data'].update('10')
                        break
        elif event == 'Start':
            form_category= values['FormCategory']
            file_OJP_path = values['form_OJB']
            file_save_path = values['form_save'] 
            number_data = int(values['number_of_data'])
            data = read_data()
            if not form_category:
                sg.Popup("Vui lòng chọn loại form ", keep_on_top=True)
            elif not file_OJP_path:
                sg.Popup('Vui lòng chọn file excel xuất của OJB', keep_on_top=True)
            elif not file_save_path:
                sg.Popup('Vui lòng chọn file form mẫu', keep_on_top=True)
            else:
                saved = False
                if form_category in data:
                    if form_category == "Y-101-Program G (Item 01+02) Shaft position KIEM TRA XUAT HANG":
                        folder_save= data[form_category].get('folder_save', '')
                        if folder_save is None or folder_save =='':
                            folder_save="C:\\"
                        base_name = os.path.basename(file_save_path)  # Ví dụ: 'test.xlsx'
                        file_name_without_extension = os.path.splitext(base_name)[0]  # Ví dụ: 'test'
                        saved = False
                        # Tạo tên file mặc định dựa trên thời gian
                        # timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%sS")
                        # default_file_name = f'{file_name_without_extension}_{timestamp}.xlsx'  # Ví dụ: 'test_20220413-080427.xlsx'
                        
                        # Tính toán tỷ lệ tương tự giữa a và b
                        base_name_ojb =os.path.basename(file_OJP_path)
                        file_name_ojb = os.path.splitext(base_name_ojb)[0]
                        similarity = difflib.SequenceMatcher(None, file_name_ojb, form_category).ratio()
                        print (file_name_without_extension,form_category,similarity)
                        while True:
                            if similarity >= 0.85:
                                # Nếu chuỗi b tương tự ít nhất 40% so với chuỗi a, thì chạy code của bạn
                                default_file_name = f'{file_name_without_extension}.xlsx'
                            
                                # Tạo đường dẫn mặc định bao gồm cả thư mục và tên file
                                default_path = os.path.join(folder_save, default_file_name)
                                default_path = default_path.replace('\\', '/')
                                # Hiển thị hộp thoại để người dùng chọn vị trí và tên file để lưu
                                # save_as = sg.popup_get_file('Lưu file', save_as=True,no_window=True,default_path=default_path, file_types=(('Excel Files', '*.xlsx'),))
                                save_as = default_path
                                
                                if file_OJP_path.endswith('.xls'):
                                    try:
                                        # Thử đọc file như là Excel file
                                        data_frame = pd.read_excel(file_OJP_path)
                                        temp_file_path = 'temp.xlsx'

                                        # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                        data_frame.to_excel(temp_file_path, index=False)

                                        # Cập nhật file_save_path để trỏ đến file tạm thời
                                        file_OJP_path = temp_file_path
                                    except ValueError:
                                        # Đọc tiêu đề của phần đầu tiên
                                        header = pd.read_csv(file_OJP_path,nrows=1, delimiter='\t',header=None)
                                        print('xong phần tiêu đề')
                                        # Đọc phần dữ liệu sau tiêu đề
                                        part1 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=1, nrows=13, header=None)
                                        print('xong phần tiêu part1')
                                        # Đọc phần dữ liệu tiếp theo sau một số dòng không xác định
                                        # Bạn sẽ cần chỉnh số skiprows để đúng với file của bạn
                                        part2 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=14, nrows=1000, header=None)
                                        print('xong phần part 2')
                                        
                                        # Kết hợp tất cả các phần lại với nhau
                                        data = pd.concat([header, part1, part2],ignore_index=True)
                                        # Tạo một Excel writer với thư viện openpyxl
                                        temp_file_path = 'temp.xlsx'
                                        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                                            data.to_excel(writer, sheet_name='Data', index=False, header=False)
                                            writer._save()
                                        print("Đã chuyển đổi xong")
                                        file_OJP_path = temp_file_path
                                    # Tạo tên file tạm thời
                                else:
                                    temp_file_path = 'temp.xlsx'

                                    # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                    data_frame.to_excel(temp_file_path, index=False)

                                    # Cập nhật file_save_path để trỏ đến file tạm thời
                                    file_OJP_path = temp_file_path
                                            
                                # Kiểm tra xem người dùng có chọn vị trí để lưu file không
                                if save_as:
                                    # Nếu có, sao chép file từ file_save_path đến vị trí được chọn
                                    if file_save_path !=save_as :
                                        print("File save path :",file_save_path,"save as : ", save_as,file_save_path !=save_as)
                                        shutil.copy2(file_save_path, save_as)
                                        file_save_path=default_path
                                    else :
                                        file_save_path=default_path
                                    if form_category in data:
                                        # Nếu có, thì lấy đoạn mã tương ứng
                                        code = textwrap.dedent(data[form_category]['code'])
                                        print("code : ", code)
                                        # Kiểm tra xem đoạn mã có tồn tại không
                                        if code and code != "":
                                            # Nếu có, thì thực thi đoạn mã
                                            exec(code)
                                        else:
                                            if form_category == "Y-101-Program G (Item 01+02) Shaft position KIEM TRA XUAT HANG":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"L15",file_save_path,"O54",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"O15",file_save_path,"T54",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101-Program G (Item 01+02) Shaft position KIEM TRA XUAT HANG', keep_on_top=True)
                                    else:
                                            if form_category == "Y-101-Program G (Item 01+02) Shaft position KIEM TRA XUAT HANG":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"L15",file_save_path,"O54",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"O15",file_save_path,"T54",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101-Program G (Item 01+02) Shaft position KIEM TRA XUAT HANG', keep_on_top=True)
                                    saved = True
                                    # if os.path.exists(temp_file_path):
                                    #     os.remove(temp_file_path)
                                    break
                            else:
                                # Nếu không, hiển thị một hộp thoại popup với hai nút
                                # layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                #         [sg.Button("Chắc chắn đúng"), sg.Button("Để kiểm tra lại")]]
                                layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                        [sg.Button("Để kiểm tra lại")]]

                                window_confirm = sg.Window("Kiểm tra nhập liệu", layout)

                                event, values = window_confirm.read()
                                window_confirm.close()

                                if event == "Để kiểm tra lại":
                                    # Nếu người dùng chọn "Để kiểm tra lại", thì thoát khỏi vòng lặp
                                    break
                                # elif event =="Chắc chắn đúng":
                                #     similarity = 0.9
                    elif form_category == "Y-101 P&G -Program 05 (Item 14+21+22+26+27+31)":
                        folder_save= data[form_category].get('folder_save', '')
                        if folder_save is None or folder_save =='':
                            folder_save="C:\\"
                        base_name = os.path.basename(file_save_path)  # Ví dụ: 'test.xlsx'
                        file_name_without_extension = os.path.splitext(base_name)[0]  # Ví dụ: 'test'
                        saved = False
                        # Tạo tên file mặc định dựa trên thời gian
                        # timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%sS")
                        # default_file_name = f'{file_name_without_extension}_{timestamp}.xlsx'  # Ví dụ: 'test_20220413-080427.xlsx'
                        
                        # Tính toán tỷ lệ tương tự giữa a và b
                        base_name_ojb =os.path.basename(file_OJP_path)
                        file_name_ojb = os.path.splitext(base_name_ojb)[0]
                        similarity = difflib.SequenceMatcher(None, file_name_ojb, form_category).ratio()
                        print (file_name_without_extension,form_category,similarity)
                        while True:
                            if similarity >= 0.85:
                                # Nếu chuỗi b tương tự ít nhất 40% so với chuỗi a, thì chạy code của bạn
                                default_file_name = f'{file_name_without_extension}.xlsx'
                            
                                # Tạo đường dẫn mặc định bao gồm cả thư mục và tên file
                                default_path = os.path.join(folder_save, default_file_name)
                                default_path = default_path.replace('\\', '/')

                                # Hiển thị hộp thoại để người dùng chọn vị trí và tên file để lưu
                                # save_as = sg.popup_get_file('Lưu file', save_as=True,no_window=True,default_path=default_path, file_types=(('Excel Files', '*.xlsx'),))
                                save_as = default_path
                                
                                if file_OJP_path.endswith('.xls'):
                                    try:
                                        # Thử đọc file như là Excel file
                                        data_frame = pd.read_excel(file_OJP_path)
                                        temp_file_path = 'temp.xlsx'

                                        # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                        data_frame.to_excel(temp_file_path, index=False)

                                        # Cập nhật file_save_path để trỏ đến file tạm thời
                                        file_OJP_path = temp_file_path
                                    except ValueError:
                                        # Đọc tiêu đề của phần đầu tiên
                                        header = pd.read_csv(file_OJP_path,nrows=1, delimiter='\t',header=None)
                                        print('xong phần tiêu đề')
                                        # Đọc phần dữ liệu sau tiêu đề
                                        part1 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=1, nrows=19, header=None)
                                        print('xong phần tiêu part1')
                                        # Đọc phần dữ liệu tiếp theo sau một số dòng không xác định
                                        # Bạn sẽ cần chỉnh số skiprows để đúng với file của bạn
                                        part2 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=20, nrows=1000, header=None)
                                        print('xong phần part 2')
                                        
                                        # Kết hợp tất cả các phần lại với nhau
                                        data = pd.concat([header, part1, part2],ignore_index=True)
                                        # Tạo một Excel writer với thư viện openpyxl
                                        temp_file_path = 'temp.xlsx'
                                        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                                            data.to_excel(writer, sheet_name='Data', index=False, header=False)
                                            writer._save()
                                        print("Đã chuyển đổi xong")
                                        file_OJP_path = temp_file_path
                                    # Tạo tên file tạm thời
                                else:
                                    temp_file_path = 'temp.xlsx'

                                    # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                    data_frame.to_excel(temp_file_path, index=False)

                                    # Cập nhật file_save_path để trỏ đến file tạm thời
                                    file_OJP_path = temp_file_path
                                            
                                # Kiểm tra xem người dùng có chọn vị trí để lưu file không
                                if save_as:
                                    # Nếu có, sao chép file từ file_save_path đến vị trí được chọn
                                    if file_save_path !=save_as :
                                        print("File save path :",file_save_path,"save as : ", save_as,file_save_path !=save_as)
                                        shutil.copy2(file_save_path, save_as)
                                        file_save_path=default_path
                                    else :
                                        file_save_path=default_path
                                    if form_category in data:
                                        # Nếu có, thì lấy đoạn mã tương ứng
                                        code = textwrap.dedent(data[form_category]['code'])
                                        print("code : ", code)
                                        # Kiểm tra xem đoạn mã có tồn tại không
                                        if code and code != "":
                                            # Nếu có, thì thực thi đoạn mã
                                            exec(code)
                                        else:
                                            if form_category == "Y-101 P&G -Program 05 (Item 14+21+22+26+27+31)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"P21",file_save_path,"AD54",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"Q21",file_save_path,"O71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"R21",file_save_path,"T71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"T21",file_save_path,"Y71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"U21",file_save_path,"Y86",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 05 (Item 14+21+22+26+27+31)', keep_on_top=True)
                                    else:
                                            if form_category == "Y-101 P&G -Program 05 (Item 14+21+22+26+27+31)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"P21",file_save_path,"AD54",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"Q21",file_save_path,"O71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"R21",file_save_path,"T71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"T21",file_save_path,"Y71",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"U21",file_save_path,"Y86",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 05 (Item 14+21+22+26+27+31)', keep_on_top=True)
                                    saved = True
                                    # if os.path.exists(temp_file_path):
                                    #     os.remove(temp_file_path)
                                    break
                            else:
                                # Nếu không, hiển thị một hộp thoại popup với hai nút
                                # layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                #         [sg.Button("Chắc chắn đúng"), sg.Button("Để kiểm tra lại")]]
                                
                                layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                        [sg.Button("Để kiểm tra lại")]]

                                window_confirm = sg.Window("Kiểm tra nhập liệu", layout)

                                event, values = window_confirm.read()
                                window_confirm.close()

                                if event == "Để kiểm tra lại":
                                    # Nếu người dùng chọn "Để kiểm tra lại", thì thoát khỏi vòng lặp
                                    break
                                # elif event =="Chắc chắn đúng":
                                #     similarity = 0.9 
                    elif form_category == "Y-101 P&G -Program 01 (Item 12+13+29)":
                        folder_save= data[form_category].get('folder_save', '')
                        if folder_save is None or folder_save =='':
                            folder_save="C:\\"
                        base_name = os.path.basename(file_save_path)  # Ví dụ: 'test.xlsx'
                        file_name_without_extension = os.path.splitext(base_name)[0]  # Ví dụ: 'test'
                        saved = False
                        # Tạo tên file mặc định dựa trên thời gian
                        # timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%sS")
                        # default_file_name = f'{file_name_without_extension}_{timestamp}.xlsx'  # Ví dụ: 'test_20220413-080427.xlsx'
                        
                        # Tính toán tỷ lệ tương tự giữa a và b
                        base_name_ojb =os.path.basename(file_OJP_path)
                        file_name_ojb = os.path.splitext(base_name_ojb)[0]
                        similarity = difflib.SequenceMatcher(None, file_name_ojb, form_category).ratio()
                        print (file_name_without_extension,form_category,similarity)
                        while True:
                            if similarity >= 0.85:
                                # Nếu chuỗi b tương tự ít nhất 40% so với chuỗi a, thì chạy code của bạn
                                default_file_name = f'{file_name_without_extension}.xlsx'
                            
                                # Tạo đường dẫn mặc định bao gồm cả thư mục và tên file
                                default_path = os.path.join(folder_save, default_file_name)
                                default_path = default_path.replace('\\', '/')

                                # Hiển thị hộp thoại để người dùng chọn vị trí và tên file để lưu
                                # save_as = sg.popup_get_file('Lưu file', save_as=True,no_window=True,default_path=default_path, file_types=(('Excel Files', '*.xlsx'),))
                                save_as = default_path
                                
                                if file_OJP_path.endswith('.xls'):
                                    try:
                                        # Thử đọc file như là Excel file
                                        data_frame = pd.read_excel(file_OJP_path)
                                        temp_file_path = 'temp.xlsx'

                                        # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                        data_frame.to_excel(temp_file_path, index=False)

                                        # Cập nhật file_save_path để trỏ đến file tạm thời
                                        file_OJP_path = temp_file_path
                                    except ValueError:
                                        # Đọc tiêu đề của phần đầu tiên
                                        header = pd.read_csv(file_OJP_path,nrows=1, delimiter='\t',header=None)
                                        print('xong phần tiêu đề')
                                        # Đọc phần dữ liệu sau tiêu đề
                                        part1 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=1, nrows=5, header=None)
                                        print('xong phần tiêu part1')
                                        # Đọc phần dữ liệu tiếp theo sau một số dòng không xác định
                                        # Bạn sẽ cần chỉnh số skiprows để đúng với file của bạn
                                        part2 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=6, nrows=1000, header=None)
                                        print('xong phần part 2')
                                        
                                        # Kết hợp tất cả các phần lại với nhau
                                        data = pd.concat([header, part1, part2],ignore_index=True)
                                        # Tạo một Excel writer với thư viện openpyxl
                                        temp_file_path = 'temp.xlsx'
                                        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                                            data.to_excel(writer, sheet_name='Data', index=False, header=False)
                                            writer._save()
                                        print("Đã chuyển đổi xong")
                                        file_OJP_path = temp_file_path
                                    # Tạo tên file tạm thời
                                else:
                                    temp_file_path = 'temp.xlsx'

                                    # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                    data_frame.to_excel(temp_file_path, index=False)

                                    # Cập nhật file_save_path để trỏ đến file tạm thời
                                    file_OJP_path = temp_file_path
                                            
                                # Kiểm tra xem người dùng có chọn vị trí để lưu file không
                                if save_as:
                                    # Nếu có, sao chép file từ file_save_path đến vị trí được chọn
                                    if file_save_path !=save_as :
                                        print("File save path :",file_save_path,"save as : ", save_as,file_save_path !=save_as)
                                        shutil.copy2(file_save_path, save_as)
                                        file_save_path=default_path
                                    else :
                                        file_save_path=default_path
                                    if form_category in data:
                                        # Nếu có, thì lấy đoạn mã tương ứng
                                        code = textwrap.dedent(data[form_category]['code'])
                                        print("code : ", code)
                                        # Kiểm tra xem đoạn mã có tồn tại không
                                        if code and code != "":
                                            # Nếu có, thì thực thi đoạn mã
                                            exec(code)
                                        else:
                                            if form_category == "Y-101 P&G -Program 01 (Item 12+13+29)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"D7",file_save_path,"AD101",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"E7",file_save_path,"C116",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"F7",file_save_path,"Y39",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 01 (Item 12+13+29)', keep_on_top=True)
                                    else:
                                            if form_category == "Y-101 P&G -Program 01 (Item 12+13+29)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"D7",file_save_path,"AD101",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"E7",file_save_path,"C116",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"F7",file_save_path,"Y39",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 01 (Item 12+13+29)', keep_on_top=True)
                                    saved = True
                                    # if os.path.exists(temp_file_path):
                                    #     os.remove(temp_file_path)
                                    break
                            else:
                                # Nếu không, hiển thị một hộp thoại popup với hai nút
                                # layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                #         [sg.Button("Chắc chắn đúng"), sg.Button("Để kiểm tra lại")]]
                                layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                        [sg.Button("Để kiểm tra lại")]]

                                window_confirm = sg.Window("Kiểm tra nhập liệu", layout)

                                event, values = window_confirm.read()
                                window_confirm.close()

                                if event == "Để kiểm tra lại":
                                    # Nếu người dùng chọn "Để kiểm tra lại", thì thoát khỏi vòng lặp
                                    break
                                # elif event =="Chắc chắn đúng":
                                #     similarity = 0.9 
                    elif form_category == "Y-101 P&G -Program 02 (Item 10+11 Left(side Welding))":
                        folder_save= data[form_category].get('folder_save', '')
                        if folder_save is None or folder_save =='':
                            folder_save="C:\\"
                        base_name = os.path.basename(file_save_path)  # Ví dụ: 'test.xlsx'
                        file_name_without_extension = os.path.splitext(base_name)[0]  # Ví dụ: 'test'
                        saved = False
                        # Tạo tên file mặc định dựa trên thời gian
                        # timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%sS")
                        # default_file_name = f'{file_name_without_extension}_{timestamp}.xlsx'  # Ví dụ: 'test_20220413-080427.xlsx'
                        
                        # Tính toán tỷ lệ tương tự giữa a và b
                        base_name_ojb =os.path.basename(file_OJP_path)
                        file_name_ojb = os.path.splitext(base_name_ojb)[0]
                        similarity = difflib.SequenceMatcher(None, file_name_ojb, form_category).ratio()
                        print (file_name_without_extension,form_category,similarity)
                        while True:
                            if similarity >= 0.85:
                                # Nếu chuỗi b tương tự ít nhất 40% so với chuỗi a, thì chạy code của bạn
                                default_file_name = f'{file_name_without_extension}.xlsx'
                            
                                # Tạo đường dẫn mặc định bao gồm cả thư mục và tên file
                                default_path = os.path.join(folder_save, default_file_name)
                                default_path = default_path.replace('\\', '/')

                                # Hiển thị hộp thoại để người dùng chọn vị trí và tên file để lưu
                                # save_as = sg.popup_get_file('Lưu file', save_as=True,no_window=True,default_path=default_path, file_types=(('Excel Files', '*.xlsx'),))
                                save_as = default_path
                                
                                if file_OJP_path.endswith('.xls'):
                                    try:
                                        # Thử đọc file như là Excel file
                                        data_frame = pd.read_excel(file_OJP_path)
                                        temp_file_path = 'temp.xlsx'

                                        # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                        data_frame.to_excel(temp_file_path, index=False)

                                        # Cập nhật file_save_path để trỏ đến file tạm thời
                                        file_OJP_path = temp_file_path
                                    except ValueError:
                                        # Đọc tiêu đề của phần đầu tiên
                                        header = pd.read_csv(file_OJP_path,nrows=1, delimiter='\t',header=None)
                                        print('xong phần tiêu đề')
                                        # Đọc phần dữ liệu sau tiêu đề
                                        part1 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=1, nrows=6, header=None)
                                        print('xong phần tiêu part1')
                                        # Đọc phần dữ liệu tiếp theo sau một số dòng không xác định
                                        # Bạn sẽ cần chỉnh số skiprows để đúng với file của bạn
                                        part2 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=7, nrows=1000, header=None)
                                        print('xong phần part 2')
                                        
                                        # Kết hợp tất cả các phần lại với nhau
                                        data = pd.concat([header, part1, part2],ignore_index=True)
                                        # Tạo một Excel writer với thư viện openpyxl
                                        temp_file_path = 'temp.xlsx'
                                        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                                            data.to_excel(writer, sheet_name='Data', index=False, header=False)
                                            writer._save()
                                        print("Đã chuyển đổi xong")
                                        file_OJP_path = temp_file_path
                                    # Tạo tên file tạm thời
                                else:
                                    temp_file_path = 'temp.xlsx'

                                    # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                    data_frame.to_excel(temp_file_path, index=False)

                                    # Cập nhật file_save_path để trỏ đến file tạm thời
                                    file_OJP_path = temp_file_path
                                            
                                # Kiểm tra xem người dùng có chọn vị trí để lưu file không
                                if save_as:
                                    # Nếu có, sao chép file từ file_save_path đến vị trí được chọn
                                    if file_save_path !=save_as :
                                        print("File save path :",file_save_path,"save as : ", save_as,file_save_path !=save_as)
                                        shutil.copy2(file_save_path, save_as)
                                        file_save_path=default_path
                                    else :
                                        file_save_path=default_path
                                    if form_category in data:
                                        # Nếu có, thì lấy đoạn mã tương ứng
                                        code = textwrap.dedent(data[form_category]['code'])
                                        print("code : ", code)
                                        # Kiểm tra xem đoạn mã có tồn tại không
                                        if code and code != "":
                                            # Nếu có, thì thực thi đoạn mã
                                            exec(code)
                                        else:
                                            if form_category == "Y-101 P&G -Program 02 (Item 10+11 Left(side Welding))":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"G8",file_save_path,"AD86",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"H8",file_save_path,"T101",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 02 (Item 10+11 Left(side Welding))', keep_on_top=True)
                                    else:
                                            if form_category == "Y-101 P&G -Program 02 (Item 10+11 Left(side Welding))":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"G8",file_save_path,"AD86",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"H8",file_save_path,"T101",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 02 (Item 10+11 Left(side Welding))', keep_on_top=True)
                                    saved = True
                                    # if os.path.exists(temp_file_path):
                                    #     os.remove(temp_file_path)
                                    break
                            else:
                                # Nếu không, hiển thị một hộp thoại popup với hai nút
                                # layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                #         [sg.Button("Chắc chắn đúng"), sg.Button("Để kiểm tra lại")]]
                                layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                        [sg.Button("Để kiểm tra lại")]]

                                window_confirm = sg.Window("Kiểm tra nhập liệu", layout)

                                event, values = window_confirm.read()
                                window_confirm.close()

                                if event == "Để kiểm tra lại":
                                    # Nếu người dùng chọn "Để kiểm tra lại", thì thoát khỏi vòng lặp
                                    break
                                # elif event =="Chắc chắn đúng":
                                #     similarity = 0.9 
                    elif form_category == "Y-101 P&G -Program 03 (10+11 Right(Side DMC)":
                        folder_save= data[form_category].get('folder_save', '')
                        if folder_save is None or folder_save =='':
                            folder_save="C:\\"
                        base_name = os.path.basename(file_save_path)  # Ví dụ: 'test.xlsx'
                        file_name_without_extension = os.path.splitext(base_name)[0]  # Ví dụ: 'test'
                        saved = False
                        # Tạo tên file mặc định dựa trên thời gian
                        # timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%sS")
                        # default_file_name = f'{file_name_without_extension}_{timestamp}.xlsx'  # Ví dụ: 'test_20220413-080427.xlsx'
                        
                        # Tính toán tỷ lệ tương tự giữa a và b
                        base_name_ojb =os.path.basename(file_OJP_path)
                        file_name_ojb = os.path.splitext(base_name_ojb)[0]
                        similarity = difflib.SequenceMatcher(None, file_name_ojb, form_category).ratio()
                        print (file_name_without_extension,form_category,similarity)
                        while True:
                            if similarity >= 0.85:
                                # Nếu chuỗi b tương tự ít nhất 40% so với chuỗi a, thì chạy code của bạn
                                default_file_name = f'{file_name_without_extension}.xlsx'
                            
                                # Tạo đường dẫn mặc định bao gồm cả thư mục và tên file
                                default_path = os.path.join(folder_save, default_file_name)
                                default_path = default_path.replace('\\', '/')

                                # Hiển thị hộp thoại để người dùng chọn vị trí và tên file để lưu
                                # save_as = sg.popup_get_file('Lưu file', save_as=True,no_window=True,default_path=default_path, file_types=(('Excel Files', '*.xlsx'),))
                                save_as = default_path
                                
                                if file_OJP_path.endswith('.xls'):
                                    try:
                                        # Thử đọc file như là Excel file
                                        data_frame = pd.read_excel(file_OJP_path)
                                        temp_file_path = 'temp.xlsx'

                                        # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                        data_frame.to_excel(temp_file_path, index=False)

                                        # Cập nhật file_save_path để trỏ đến file tạm thời
                                        file_OJP_path = temp_file_path
                                    except ValueError:
                                        # Đọc tiêu đề của phần đầu tiên
                                        header = pd.read_csv(file_OJP_path,nrows=1, delimiter='\t',header=None)
                                        print('xong phần tiêu đề')
                                        # Đọc phần dữ liệu sau tiêu đề
                                        part1 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=1, nrows=6, header=None)
                                        print('xong phần tiêu part1')
                                        # Đọc phần dữ liệu tiếp theo sau một số dòng không xác định
                                        # Bạn sẽ cần chỉnh số skiprows để đúng với file của bạn
                                        part2 = pd.read_csv(file_OJP_path, delimiter='\t', skiprows=7, nrows=1000, header=None)
                                        print('xong phần part 2')
                                        
                                        # Kết hợp tất cả các phần lại với nhau
                                        data = pd.concat([header, part1, part2],ignore_index=True)
                                        # Tạo một Excel writer với thư viện openpyxl
                                        temp_file_path = 'temp.xlsx'
                                        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                                            data.to_excel(writer, sheet_name='Data', index=False, header=False)
                                            writer._save()
                                        print("Đã chuyển đổi xong")
                                        file_OJP_path = temp_file_path
                                    # Tạo tên file tạm thời
                                else:
                                    temp_file_path = 'temp.xlsx'

                                    # Ghi dữ liệu vào file tạm thời với định dạng .xlsx
                                    data_frame.to_excel(temp_file_path, index=False)

                                    # Cập nhật file_save_path để trỏ đến file tạm thời
                                    file_OJP_path = temp_file_path
                                            
                                # Kiểm tra xem người dùng có chọn vị trí để lưu file không
                                if save_as:
                                    # Nếu có, sao chép file từ file_save_path đến vị trí được chọn
                                    if file_save_path !=save_as :
                                        print("File save path :",file_save_path,"save as : ", save_as,file_save_path !=save_as)
                                        shutil.copy2(file_save_path, save_as)
                                        file_save_path=default_path
                                    else :
                                        file_save_path=default_path
                                    if form_category in data:
                                        # Nếu có, thì lấy đoạn mã tương ứng
                                        code = textwrap.dedent(data[form_category]['code'])
                                        print("code : ", code)
                                        # Kiểm tra xem đoạn mã có tồn tại không
                                        if code and code != "":
                                            # Nếu có, thì thực thi đoạn mã
                                            exec(code)
                                        else:
                                            if form_category == "Y-101 P&G -Program 03 (10+11 Right(Side DMC)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"G8",file_save_path,"O101",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"H8",file_save_path,"Y101",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 03 (10+11 Right(Side DMC)', keep_on_top=True)
                                    else:
                                            if form_category == "Y-101 P&G -Program 03 (10+11 Right(Side DMC)":
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"G8",file_save_path,"O101",number_data)
                                                copy_from_excel_to_excel_vertical(file_OJP_path,"H8",file_save_path,"Y101",number_data)
                                                sg.Popup('Hoàn thành copy dữ liệu từ file Y-101 P&G -Program 03 (10+11 Right(Side DMC)', keep_on_top=True)
                                    saved = True
                                    # if os.path.exists(temp_file_path):
                                    #     os.remove(temp_file_path)
                                    break
                            else:
                                # Nếu không, hiển thị một hộp thoại popup với hai nút
                                # layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                #         [sg.Button("Chắc chắn đúng"), sg.Button("Để kiểm tra lại")]]
                                
                                layout = [[sg.Text("Hình như bạn đang nhập nhầm chương trình thì phải")],
                                        [sg.Button("Để kiểm tra lại")]]

                                window_confirm = sg.Window("Kiểm tra nhập liệu", layout)

                                event, values = window_confirm.read()
                                window_confirm.close()

                                if event == "Để kiểm tra lại":
                                    # Nếu người dùng chọn "Để kiểm tra lại", thì thoát khỏi vòng lặp
                                    break
                                # elif event =="Chắc chắn đúng":
                                #     similarity = 0.9 
                else:
                    sg.Popup('Khả năng cao chương trình có vấn đề rồi đó, liên lạc người viết ra chương trình này nhé. ', keep_on_top=True)
        elif event == 'FormCategory':
            data = read_data()
            form_category = values['FormCategory']
            if form_category in data:
                if isinstance(data[form_category], dict):  # Kiểm tra xem data[form_category] có phải là một từ điển hay không
                    window['form_OJB'].update(data[form_category].get('form_OJB', ''))
                    window['form_save'].update(data[form_category].get('form_save', ''))
                    if data[form_category].get('number_of_data','') is None or data[form_category].get('number_of_data','') == '':
                        window['number_of_data'].update('10')
                    else:
                        window['number_of_data'].update(data[form_category].get('number_of_data',''))
                else:
                    window['form_OJB'].update('')
                    window['form_save'].update('')
                    window['number_of_data'].update('10')
            else:
                window['form_OJB'].update('')
                window['form_save'].update('')
                window['number_of_data'].update('10')
        elif event == "Open":
            try :
                if saved:
                    # Thực hiện các hành động "Open" ở đây
                    # Ví dụ: mở file vừa được lưu
                    print("Open file",file_save_path)
                    os.startfile(file_save_path)
                    # saved = False
                    # Sự kiện khi người dùng nhấn nút 'Update'
                else:
                   sg.popup("Vui lòng thực hiện bước 'Start' và 'Save As' trước khi mở file.", keep_on_top=True)
            except UnboundLocalError:
                 sg.popup("Vui lòng thực hiện bước 'Start' và 'Save As' trước khi mở file.", keep_on_top=True)
        elif event == 'Close':
            break
            
    windowOGP.close()

