import csv
from datetime import datetime
import os
import openpyxl
import serial
import PySimpleGUI as sg



def create_excel_1():

    # folder_path = 'C:/may1'
    folder_path = values['folder_save']
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = f'data_{timestamp}.xlsx'

    file_path1 = os.path.join(folder_path, file_name)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_titles = ["Maximum amplitude [mm]", "Drive frequency(Δmax) [Hz]", "Res frequency [Hz]",
                     "Res frequency range [Hz]", "Damping factor", "Back-EMF factor [V/mm]", "Motor constant [V/(m/s)]"]

    for col_num, title in enumerate(column_titles, start=1):
        sheet.cell(row=1, column=col_num, value=title)

        sheet.column_dimensions[chr(64 + col_num)].width = 25 

        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path1)
    return file_path1


def append_to_excel_1(data,file_path1):
    workbook = openpyxl.load_workbook(file_path1)
    sheet = workbook.active

    sheet.append(data)

    for cell in sheet[sheet.max_row]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path1)


def create_excel_2():
    # folder_path = 'C:/may2'
    folder_path = values['folder_save']

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    file_name = f'data_{timestamp}.xlsx'

    file_path2 = os.path.join(folder_path, file_name)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_titles = ["Tần số cộng hưởng Fr [Hz]", "Hệ số suy giảm D"]

    for col_num, title in enumerate(column_titles, start=1):
        sheet.cell(row=1, column=col_num, value=title)
        
        sheet.column_dimensions[chr(64 + col_num)].width = 25 

        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path2)
    return file_path2


def append_to_excel_2(data,file_path2):
    workbook = openpyxl.load_workbook(file_path2)
    sheet = workbook.active

    sheet.append(data)

    for cell in sheet[sheet.max_row]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path2)



# def main():
layout = [
    [sg.Text("Chọn folder lưu file: ",font=('Helvetica', 20))],[sg.InputText(disabled=True,size=(30, 1),font=('Helvetica', 20), key='location_save'),sg.FolderBrowse(font=('Helvetica', 15),key='folder_save')],
    [sg.Text("                      Máy kiểm tra",font=('Helvetica', 20), justification='center', pad=(20, 20),key='text_title', text_color='yellow',enable_events=True)],
    # [sg.Text("",font=('Helvetica', 20),key='text_title')],

    [sg.Button('Máy 1', size=(15, 2), button_color=('white', 'green'), font=('Helvetica', 16), pad=(40, 40),enable_events=True),
        sg.Button('Máy 2', size=(15, 2), button_color=('white', 'blue'), font=('Helvetica', 16), pad=(40, 40),enable_events=True)]
]


window = sg.Window('My GUI', layout, resizable=False, finalize=True)



location_save = "C:/"
with open('save.txt','r') as f:
    l_save = f.read()
window['location_save'].update(value =l_save )
try:
    while True:
        event, values = window.read(timeout=20)
        try:
            if values['location_save'] != "":
                location_save = values['location_save']
        except:
            pass
        if event == sg.WINDOW_CLOSED:
            print(location_save)
            with open('save.txt','w') as f:
                f.write(location_save)
            break

        elif event == 'Máy 1':

            i = 1
            max_com_ports = 100
            success = False

            while i <= max_com_ports:
                port = f'COM{i}'
                try:
                    ser = serial.Serial(port, 38400, timeout=1)
                    print(f'Successfully opened {port}')
                    success = True
                    break 
                except serial.SerialException:
                    print(f'Failed to open {port}')
                
                i += 1

            if not success:
                print("Failed to open any COM port after 100 attempts.")
                sg.popup_error('Connection error', title='Notification', background_color='lightblue', text_color='red',
                        no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))

            else:
                f1 = 1
                try:

                    while True:
                        event, values = window.read(timeout=20)
                        try:
                            if values['location_save'] != "":
                                location_save = values['location_save']
                        except:
                            pass
                        window['Máy 1'].update(disabled= True)
                        window['Máy 2'].update(disabled= True)

                        window['text_title'].update('               Đang đo máy 1')

                        received_data1 = ser.readline().decode('ascii').strip()
                        
                        data_elements1 = received_data1.split(',')
                        if data_elements1 != ['']:
                            if f1 == 1:
                                file_path1 = create_excel_1()
                                f1+=1

                            relevant_data1 = data_elements1[8:15]  
                            append_to_excel_1(relevant_data1,file_path1)
                        else:
                            print('data1: None')
                        if event == sg.WINDOW_CLOSED:
                            break




                finally:
                    with open('save.txt','w') as f:
                        f.write(location_save)
                    ser.close()
                    sg.popup('Done', title='Notification', background_color='lightblue', text_color='green',
                        no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))


        elif event == 'Máy 2':
            i = 1
            max_com_ports = 100
            success = False

            while i <= max_com_ports:
                port = f'COM{i}'
                try:
                    ser = serial.Serial(port, 38400, timeout=1)
                    print(f'Successfully opened {port}')
                    success = True
                    break 
                except serial.SerialException:
                    print(f'Failed to open {port}')
                
                i += 1

            if not success:
                print("Failed to open any COM port after 100 attempts.")
                sg.popup_error('Connection error', title='Notification', background_color='lightblue', text_color='red',
                        no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))

            else:
                f2 = 1
                try:

                    while True:
                        event, values = window.read(timeout=20)
                        try:
                            if values['location_save'] != "":
                                location_save = values['location_save']
                        except:
                            pass
                        window['Máy 1'].update(disabled= True)
                        window['Máy 2'].update(disabled= True)
                        window['text_title'].update('               Đang đo máy 2')

                        received_data2 = ser.readline().decode('ascii').strip()
                        data_elements2 = received_data2.split(',')
                        if data_elements2 != ['']:
                            if f2 ==1:
                                file_path2 = create_excel_2()
                                f2+=1

                            relevant_data2 = data_elements2[7:9]  
                            append_to_excel_2(relevant_data2,file_path2)
                        else:
                            print('data2: None')

                        if event == sg.WINDOW_CLOSED:

                            break


                finally:
                    with open('save.txt','w') as f:
                        f.write(location_save)
                    ser.close()
                    sg.popup('Done', title='Notification', background_color='lightblue', text_color='green',
                        no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))



    window.close()
except:
    pass


#pyinstaller --onefile --noconsole connect.py