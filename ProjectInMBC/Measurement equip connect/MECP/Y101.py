#Khai báo thư viện:
from datetime import datetime # Lib ngày giờ hệ thống
import os       # Lib xử lý tập tin, thư mục ....liên quan hệ điều hành
import openpyxl  # Lib xử lý file excel
import serial    # Lib giao tiếp cổng com
import PySimpleGUI as sg    # Lib giao diện
import lib

# Các funtion:
# Hàm tạo file excel dùng cho máy đo actuator
def create_excel_1():
    folder_path = location_save_var
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = f'data_Actuator_{timestamp}.xlsx'

    file_path1 = os.path.join(folder_path, file_name)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    

    column_titles = ["Maximum amplitude [mm]", "Drive frequency(Δmax) [Hz]", "Res frequency [Hz]",
                     "Res frequency range [Hz]", "Damping factor", "Back-EMF factor [V/mm]", "Motor constant [V/(m/s)]"]

    # ghi hạng mục đo vào file excel actuator
    #for col_num, title in enumerate(column_titles, start=1):
    #    sheet.cell(row=1, column=col_num, value=title)
    j=1  
    for col_num in column_titles:
        sheet.cell(row=1, column=j,value= col_num)
        j+=1
        #sheet.column_dimensions[chr(64 + col_num)].width = 25 
        
        #cell = sheet.cell(row=1, column=col_num)
        #cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
    workbook.save(file_path1)
    return file_path1

# Hàm đưa dữ liệu vào fiel excel
def append_to_excel_1(data,file_path1):
    """
    Đưa dữ liệu vào file excel
    """
    workbook = openpyxl.load_workbook(file_path1)
    sheet = workbook.active
    sheet.append(data)

    for cell in sheet[sheet.max_row]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path1)

# Ghi dữ liệu Actuator vào form ghi chép thành phẩm
def append_to_form_ac(data,form_path,count):
    wb = openpyxl.load_workbook(form_path)
    sheet=wb['SAMPLING_INSPECTION_EN']

    # 4. Actuator damping factor    
    sheet[f'AD{8 + count}'] = float(data[4])
    # 5. Actuator frenquency range
    sheet[f'O{23 + count}'] = float(data[3])
    # 10. Motor constant
    sheet[f'T{38 + count}'] = float(data[6])
    # Res frequency
    sheet[f'C{100 + count}'] = float(data[2])

    #file_name=os.path.basename(form_path) # lấy tên file từ đường dẫn
    wb.save(form_path)  


#Ghi dữ liệu Resonator vào form ghi chép thành phẩm:
def append_to_form_re(data,form_path,count):
    wb = openpyxl.load_workbook(form_path)
    sheet=wb['SAMPLING_INSPECTION_EN']

    # 1. Resonator requence    
    sheet[f'O{8 + count}'] = float(data[0])
    # 5. Actuator frenquency range
    sheet[f'T{8 + count}'] = float(data[1])
    
    #file_name=os.path.basename(form_path) # lấy tên file từ đường dẫn
    wb.save(form_path)      

# Ghi dữ liệu AC vào form tuổi thọ:
def append_to_form_ac_tt(data,form_path,time_cell, count):
    wb = openpyxl.load_workbook(form_path)
    sheet = wb['data']
    if time_cell == '0h':
        # 4. Actuator damping factor    
        sheet[f'D{4 + count}'] = float(data[4])
        # 5. Actuator frenquency range
        sheet[f'C{4 + count}'] = float(data[3])
        # 10. Motor constant
        sheet[f'E{4 + count}'] = float(data[6])
        # Res frequency
        sheet[f'B{4 + count}'] = float(data[2])
    elif time_cell == '100h':
         # 4. Actuator damping factor    
        sheet[f'J{4 + count}'] = float(data[4])
        # 5. Actuator frenquency range
        sheet[f'I{4 + count}'] = float(data[3])
        # 10. Motor constant
        sheet[f'K{4 + count}'] = float(data[6])
        # Res frequency
        sheet[f'H{4 + count}'] = float(data[2])
    elif time_cell == '200h':
        # 4. Actuator damping factor    
        sheet[f'P{4 + count}'] = float(data[4])
        # 5. Actuator frenquency range
        sheet[f'O{4 + count}'] = float(data[3])
        # 10. Motor constant
        sheet[f'Q{4 + count}'] = float(data[6])
        # Res frequency
        sheet[f'N{4 + count}'] = float(data[2])
    elif time_cell == '242h':
         # 4. Actuator damping factor    
        sheet[f'V{4 + count}'] = float(data[4])
        # 5. Actuator frenquency range
        sheet[f'U{4 + count}'] = float(data[3])
        # 10. Motor constant
        sheet[f'W{4 + count}'] = float(data[6])
        # Res frequency
        sheet[f'T{4 + count}'] = float(data[2])

    wb.save(form_path)
    
# Ghi dữ liệu RE vào form tuổi thọ
def append_to_form_re_tt(data,form_path,time_cell, count):
    wb = openpyxl.load_workbook(form_path)
    sheet = wb['data']
    if time_cell == '0h':
        # 1. Resonator requence    
        sheet[f'F{4 + count}'] = float(data[0])
        # 5. Actuator frenquency range
        sheet[f'G{4 + count}'] = float(data[1])
    elif time_cell == '100h':
        # 1. Resonator requence    
        sheet[f'L{4 + count}'] = float(data[0])
        # 5. Actuator frenquency range
        sheet[f'M{4 + count}'] = float(data[1])
    elif time_cell == '200h':
        # 1. Resonator requence    
        sheet[f'R{4 + count}'] = float(data[0])
        # 5. Actuator frenquency range
        sheet[f'S{4 + count}'] = float(data[1])
    elif time_cell == '242h':
        # 1. Resonator requence    
        sheet[f'X{4 + count}'] = float(data[0])
        # 5. Actuator frenquency range
        sheet[f'Y{4 + count}'] = float(data[1])

    wb.save(form_path)
# Hàm tạo fiel excel dùng cho máy đo resonator
def create_excel_2():
    
    folder_path = location_save_var

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    file_name = f'data_Resonator_{timestamp}.xlsx'

    file_path2 = os.path.join(folder_path, file_name)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_titles = ["Resonator Frequency Fr [Hz]", "Resonator Damping Factor D"]
    
    #ghi hạng mục vào file excel
    
    #for col_num, title in enumerate(column_titles, start=1):
    #    sheet.cell(row=1, column=col_num, value=title)
    n=1
    for col_num in column_titles:
        sheet.cell(row=1, column=n, value= col_num)
        n+=1    
        #sheet.column_dimensions[chr(64 + col_num)].width = 25 

        #cell = sheet.cell(row=1, column=col_num)
        #cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path2)
    return file_path2

# Hàm đưa dữ liệu vào file excel2
def append_to_excel_2(data,file_path2):
    workbook = openpyxl.load_workbook(file_path2)
    sheet = workbook.active

    sheet.append(data)

    for cell in sheet[sheet.max_row]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(file_path2)


# Liên quan Form

def main():
    sg.theme_text_color('black')
    layout = [
        [sg.Text("Chọn folder lưu file: ")],
        [sg.InputText(disabled=True,size=(50, 1), key='location_save'),sg.FolderBrowse(key='folder_save'),sg.Button("Update")],
        [sg.Text("Chọn form ghi chép thành phẩm: ")],
        [sg.InputText(size=(50, 1), key='form_save'),sg.FilesBrowse(key='form_save_br'), sg.Button('Clear')],
        [sg.Text("Chọn form ghi chép Thí nghiệm tuổi thọ: ")],
        [sg.InputText(size=(50, 1), key='form_save_tt'),sg.FilesBrowse(key='form_save_br_tt'), sg.Button('Clear', key='Clear_tt'),sg.Text('Time:') ,sg.Combo(values=['0h', '100h','200h','242h'], key='Time_cell')],
        [sg.Text('Status:'),sg.Text("",font=('Helvetica', 13,'bold'), justification='center', pad=(20, 20),key='text_title', text_color='yellow',enable_events=True, expand_x=True),],
        
        [sg.Frame('Actuator',[
            [sg.Text('Maximum amplitude [mm]'),sg.InputText(size=10,key='data1')],
            [sg.Text('Drive frequency(Δmax)[Hz]'),sg.InputText(size=10,key='data2')],
            [sg.Text('Res frequency [Hz]          '),sg.InputText(size=10,key='data3')],
            [sg.Text('Res frequency range[Hz]   '),sg.InputText(size=10,key='data4')],
            [sg.Text('Damping factor                '),sg.InputText(size=10,key='data5')],
            [sg.Text('Back-EMF factor[V/mm]   '),sg.InputText(size=10,key='data6')],
            [sg.Text('Motor constant [V/(m/s)]  '),sg.InputText(size=10,key='data7')]
         ]),
         sg.Frame('Resonator',[
            [sg.Text('Resonator Frequency Fr [Hz]'),sg.InputText(size=10,key='data8')],
            [sg.Text('Resonator Damping Factor D '),sg.InputText(size=10,key='data9')]
         ],expand_y=True)   
        ],
        
        [sg.Button('Actuator', size=(15, 2), button_color=('white', 'green'), font=('Helvetica', 16,'bold'), pad=(40, 40),enable_events=True),
            sg.Button('Stop', size=(10,2),button_color=('white','red'),enable_events=True,font=('Helvetica',10,'bold')),
            sg.Button('Resonator', size=(15, 2), button_color=('white', 'blue'), font=('Helvetica', 16,'bold'), pad=(40, 40),enable_events=True)],

        [sg.Text("VDM-Inspection Section 2023/11",font=('Helvetica',8))]    
    ]
    window = sg.Window('Measurement equip connect Programe - Y101 Performance', layout, resizable=False, finalize=True)    
    # Kiểm tra file save.txt có tồn tại hay không, nếu không tồn tại thì tạo ra file save.txt tại đường dẫn local
    if not os.path.isfile('save.txt'):
        location_save = "C:/"
        with open('save.txt','w') as f:
            f.write(location_save)
        window['location_save'].update(value=location_save) # hiển thị đường dẫn ra form

    # Nếu file save.txt đã tồn tại thì lấy nội dung trong file save.txt vào biến location_save
    else: 
        with open('save.txt','r') as f:
            l_save = f.read()
        window['location_save'].update(value =l_save )
    
    try:
        
        while True:
            event, values = window.read(timeout=20)
          
            if values['location_save'] != "":
                
                global location_save_var
                global form_save_var
                location_save_var=values['location_save']
                form_save_var= values['form_save']
                form_save_tt_var= values['form_save_tt']
                
                if event == sg.WIN_CLOSED:
                    break

                elif event == 'Update':
                    f= open('save.txt','w')
                    f.write(location_save_var)
                    sg.popup('Completed Update đường dẫn lưu dữ liệu',title='Completed')
                elif event == 'Clear' or event == 'Clear_tt':
                    window['form_save'].update('')
                    window['form_save_tt'].update('')
                    
                # Start measure for Actuator equip
                elif event == 'Actuator':
                    i = 1
                    max_com_ports = 100
                    success = False
                    
                    #Khai báo cổng com
                    while i <= max_com_ports:
                        port = f'COM{i}'
                        try:
                            ser = serial.Serial(port, 38400, timeout=1)
                            print(f'Successfully opened {port}')
                            success = True
                            break 
                        except serial.SerialException:
                            print(f'Failed to open {port}')
                        i += 1
                    
                    if not success:
                        print("Failed to open any COM port after 100 attempts.")
                        sg.popup_error('Connection error', title='Notification', background_color='lightblue', text_color='red',
                                no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))

                    else:
                        f1 = 1
                        if form_save_var != "":
                           ask= lib.input_box('Bạn muốn đo từ con MTR thứ:')
                           i= int(ask) 
                        
                        # nạp dữ liệu biến cần đo khi đo thí nghiệm tuổi thọ
                       
                        if form_save_tt_var != "" :
                            
                            i1=1
                            i2=1
                            ask = lib.input_box('Bạn muốn đo từ con MTR thứ (tuổi thọ):')
                            i2= 3 * int(ask) - 2 # Hàm tính toán vị trí ô cell nhập dữ liệu, vì mỗi con đo ba lần
                            time_cell = values['Time_cell']
                            if time_cell == "" :
                                time_cell = "0h"
                           
                        try:

                            while True:
                                event, values = window.read(timeout=20)
                                
                                window['Actuator'].update(disabled= True)
                                window['Resonator'].update(disabled= True)
                                window['Update'].update(disabled = True)
                                window['Clear'].update(disabled=True)
                                window['Clear_tt'].update(disabled=True)
                                window['folder_save'].update(disabled=True)
                                window['form_save_br'].update(disabled=True)
                                window['form_save_br_tt'].update(disabled=True)
                                window['form_save_tt'].update(disabled=True)

                                window['text_title'].update('Đang đo Actuator')

                                received_data1 = ser.readline().decode('ascii').strip()
                                
                                data_elements1 = received_data1.split(',')
                                if event == sg.WINDOW_CLOSED or event == 'Stop':
                                    re_form()
                                    break

                                elif data_elements1 != ['']:
                                    
                                    if f1 == 1:
                                        file_path1 = create_excel_1()
                                        f1+=1
                                        
                                    # Lưu dữ liệu Actuator vào file excel
                                    relevant_data1 = data_elements1[8:15]  
                                    append_to_excel_1(relevant_data1,file_path1)
                                    # Hiển thị data Actuator lên formt
                                    
                                    window['data1'].update(data_elements1[8])
                                    window['data2'].update(data_elements1[9])
                                    window['data3'].update(data_elements1[10])
                                    window['data4'].update(data_elements1[11])
                                    window['data5'].update(data_elements1[12])
                                    window['data6'].update(data_elements1[13])
                                    window['data7'].update(data_elements1[14])

                                    # lưu dữ liệu vào form ghi chép thành phẩm
                                    if form_save_var != "" and form_save_tt_var =="":
                                        if i< 11 :
                                            ask = sg.popup_yes_no('Lưu dữ liệu Actuator vào form ghi chép',f'Sample No: {i}',title='Thông báo')
                                            if ask == 'Yes' :
                                                append_to_form_ac(relevant_data1,form_save_var,i)
                                                i+=1
                                        else:
                                            sg.popup_ok('Đã đủ 10pcs dữ liệu', 'Không thể lưu thêm' )

                                    # Đưa dữ liệu ra form tuổi thọ        
                                    elif form_save_tt_var != "" and form_save_var == "":
                                        if i1 < 4 and i2 < 61:
                                            ask = sg.popup_yes_no('Lưu dữ liệu Actuator vào form', f'Lần đo thứ:{i1}', title="Thông báo")
                                            if ask == 'Yes':
                                                append_to_form_ac_tt(relevant_data1,form_save_tt_var,time_cell,i2)
                                                i3=i2//3
                                                i1 +=1
                                                i2 +=1
                                                
                                                if i1 == 4 :
                                                    ask = sg.popup_yes_no(f'Đã đo 3 lần cho MTR NO: {i3}, bạn muốn đo MTR tiếp theo:', title='Thông báo')
                                                    if ask == 'Yes':
                                                        i1=1
                                                
                                                if i2 == 61:
                                                    ask = sg.popup_ok(' Bạn đã đo 20 pcs, bạn không thể tiếp tục đo,', ' do vượt quá phạm vi ghi chép của form', title='Thông báo')
                                        else:
                                            sg.popup_ok('Vượt quá số lần đo', title='Thông báo')        
                                else:
                                    print('Data Actuator: None')
                        
                        finally:
                            ser.close()
                            sg.popup('Compeleted', title='Notification',font=('Helvetica',16))

                # Start measure for Resonator equip
                elif event == 'Resonator':
                    i = 1
                    max_com_ports = 100
                    success = False

                    while i <= max_com_ports:
                        port = f'COM{i}'
                        try:
                            ser = serial.Serial(port, 38400, timeout=1)
                            print(f'Successfully opened {port}')
                            success = True
                            break 
                        except serial.SerialException:
                            print(f'Failed to open {port}')
                        
                        i += 1
                    if not success:
                   
                        print("Failed to open any COM port after 100 attempts.")
                        sg.popup_error('Connection error', title='Notification', background_color='lightblue', text_color='red',
                                no_titlebar=True, grab_anywhere=True,font=('Helvetica',18))
                        
                    else:
                        f2 = 1
                        if form_save_var != "":
                            ask = lib.input_box("Bạn muốn đo từ con MTR thứ:")
                            i=int(ask)
                        
                        if form_save_tt_var != "" :
                            i1=1
                            i2=1
                            ask = lib.input_box('Bạn muốn đo từ con MTR thứ (tuổi thọ):')
                            i2= 3 * int(ask) - 2  # Hàm tính toán ô cell đưa dữ liệu vào vì một con đo 3 lần
                            time_cell = values['Time_cell']
                            if time_cell == "" :
                                time_cell = "0h"

                        try:

                            while True:
                                event, values = window.read(timeout=20)
                                
                                window['Actuator'].update(disabled= True)
                                window['Resonator'].update(disabled= True)
                                window['Update'].update(disabled = True)
                                window['Clear'].update(disabled=True)
                                window['Clear_tt'].update(disabled=True)
                                window['folder_save'].update(disabled=True)
                                window['form_save_br'].update(disabled=True)
                                window['form_save_tt'].update(disabled=True)
                                window['form_save_br_tt'].update(disabled=True)

                                window['text_title'].update('Đang đo Resonator')

                                received_data2 = ser.readline().decode('ascii').strip()
                                data_elements2 = received_data2.split(',')

                                if event == sg.WINDOW_CLOSED or event == 'Stop':
                                    re_form()
                                    break
                                elif data_elements2 != ['']:
                                    if f2 ==1:
                                        file_path2 = create_excel_2()
                                        f2+=1

                                    # Lưu data Resonator vào file excel
                                    relevant_data2 = data_elements2[7:9]  
                                    append_to_excel_2(relevant_data2,file_path2)
                                    # hiển thị data Resonator ra form
                                    window['data8'].update(data_elements2[7])
                                    window['data9'].update(data_elements2[8])

                                    # lưu dữ liệu vào form ghi chép
                                    if form_save_var != "" and form_save_tt_var == "":
                                        if i< 11 :
                                            ask = sg.popup_yes_no('Lưu dữ liệu Resonator vào form ghi chép',f'Sample No: {i}',title='Thông báo')
                                            if ask == 'Yes' :
                                                append_to_form_re(relevant_data2,form_save_var,i)
                                                i+=1
                                        else:
                                            sg.popup_ok('Đã đủ 10pcs dữ liệu', 'Không thể lưu thêm' )
                                    elif form_save_tt_var != "" and form_save_var =="" :
                                        if i1 < 4 and i2 < 61:
                                            ask = sg.popup_yes_no('Lưu dữ liệu Resonator vào form', f'Lần đo thứ:{i1}', title="Thông báo")
                                            if ask == 'Yes':
                                                append_to_form_re_tt(relevant_data2,form_save_tt_var,time_cell,i2)
                                                i3= i2//3
                                                i1 +=1
                                                i2 +=1
                                                if i1 == 4 :
                                                    ask = sg.popup_yes_no(f'Đã đo 3 lần cho MTR NO: {i3}, bạn muốn đo MTR tiếp theo:', title='Thông báo')
                                                    if ask == 'Yes':
                                                        i1=1
                                                
                                                if i2 == 61:
                                                    ask = sg.popup_ok(' Bạn đã đo 20 pcs, bạn không thể tiếp tục đo,', ' do vượt quá phạm vi ghi chép của form', title='Thông báo')
                                        else:
                                            sg.popup_ok('Vượt quá số lần đo', title='Thông báo')   

                                else:
                                    print('Data Resonator: None')
                                    

                        finally:
                            ser.close()
                            sg.popup('Compeleted', title='Notification',font=('Helvetica',16))
                # reset form về trạng thái ban đầu    
                def re_form():
                    window['Actuator'].update(disabled= False)
                    window['Resonator'].update(disabled= False)
                    window['Update'].update(disabled = False)
                    window['Clear'].update(disabled=False)
                    window['Clear_tt'].update(disabled=False)
                    window['folder_save'].update(disabled=False)
                    window['form_save_br'].update(disabled=False)
                    window['form_save_br_tt'].update(disabled=False)
                    window['form_save_tt'].update(disabled=False)

                    window['text_title'].update('')   
                  
            else:
                msgbox=sg.popup("Chưa có đường dẫn lưu file",button_type=0)
                
        window.close()
    except:
        pass



#pyinstaller --onefile --noconsole connect.py