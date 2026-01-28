import PySimpleGUI as sg
import json
import openpyxl  # Lib xử lý file excel
from datetime import datetime # Lib ngày giờ hệ thống
import os 

def input_box(title):
    """
    title: Nhập tiêu đề Input box
    """ 
    layout =[
        [sg.Text('', key='-content_t-')],
        [sg.Input(key='-content-' )],
        [sg.Button('OK', bind_return_key=True)]
    ]
    window = sg.Window('Input box', layout, resizable=True,finalize=True)

    while True:
        events, values = window.read(timeout=20)
        window['-content_t-'].update(title)
        if events == sg.WINDOW_CLOSED :
            break
        elif events == 'OK' :
            value = values['-content-']
            if value == "" :
                sg.popup_ok('Bạn phải nhập dữ liệu', title= 'Thông báo')
            else:    
                break
    
    window.close()
    return value

# input_box("Check Input")

# Create List and checkbox
def create_Box():
        # Load data from data.txt if it exists
    try:
        with open('data.txt', 'r') as f:
            saved_data = json.load(f)
    except FileNotFoundError:
        saved_data = {'selected_items': [], 'checked_options': []}

    # Define the layout
    layout = [
    [sg.Listbox(values=['ALL', 'Item 1', 'Item 2', 'Item 3'], select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE, key='-LIST-', default_values=saved_data['selected_items'], size=(20, 10))],
    [sg.Checkbox('ALL', key='-ALL-', default=('ALL' in saved_data['checked_options']))],
    [sg.Checkbox('Option 1', key='Option 1', default=('Option 1' in saved_data['checked_options'])),
     sg.Checkbox('Option 2', key='Option 2', default=('Option 2' in saved_data['checked_options'])),
     sg.Checkbox('Option 3', key='Option 3', default=('Option 3' in saved_data['checked_options']))],
    [sg.Button('Submit')]
]

    window = sg.Window('List and Checkbox Frames', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        elif event == '-ALL-':
            window['Option 1'].update(value=values['-ALL-'])
            window['Option 2'].update(value=values['-ALL-'])
            window['Option 3'].update(value=values['-ALL-'])
        elif event == '-LIST-':
            if 'ALL' in values['-LIST-']:
                window['-LIST-'].update(set_to_index=[i for i in range(4)])
            else:
                window['-LIST-'].update(set_to_index=[i for i, item in enumerate(window['-LIST-'].get_list_values()) if item in values['-LIST-']])
        elif event == 'Submit':
            selected_items = values['-LIST-']
            checked_options = [key for key, checked in values.items() if checked and key != '-LIST-']
            
            # Save the selected items and checked options to data.txt
            with open('data.txt', 'w') as f:
                json.dump({'selected_items': selected_items, 'checked_options': checked_options}, f)
            
            # Create layout for the new window
            new_layout = [
                [sg.Text('Selected items:')],
                [sg.Text(item) for item in selected_items],
                [sg.Text('Checked options:')],
                [sg.Text(option) for option in checked_options],
                [sg.Button('Close')]
            ]
            
            # Create and display the new window
            new_window = sg.Window('Selected Items and Checked Options', new_layout)
            while True:
                new_event, new_values = new_window.read()
                if new_event == sg.WINDOW_CLOSED or new_event == 'Close':
                    new_window.close()
                    break

    window.close()
    
# create_Box()

def create_excel():
    folder_path = "C:\\Users\\12953 bao\\Desktop\\"
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamp = datetime.now().strftime("%Y-%m-%d %H_%M_%S")
    file_name = f'data_Actuator_{timestamp}.xlsx'

    file_path1 = os.path.join(folder_path, file_name)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    

    column_titles = ["Maximum amplitude [mm]", "Drive frequency(Δmax) [Hz]", "Res frequency [Hz]",
                     "Res frequency range [Hz]", "Damping factor", "Back-EMF factor [V/mm]", "Motor constant [V/(m/s)]"]

    # ghi hạng mục đo vào file excel actuator
    #for col_num, title in enumerate(column_titles, start=1):
    #    sheet.cell(row=1, column=col_num, value=title)
    j=1  
    for col_num in column_titles:
        sheet.cell(row=1, column=j,value= col_num)
        j+=1
        #sheet.column_dimensions[chr(64 + col_num)].width = 25 
        
        #cell = sheet.cell(row=1, column=col_num)
        #cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
    workbook.save(file_path1)
    return file_path1

# create_excel()

