import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import os
import pandas as pd
import calendar
import datetime
from ult.FileMontlyData.Guidle import stateMontly
import openpyxl
import json
import xlwings as xw
import zipfile

CHECK_DIR = os.path.join(os.getcwd(), "DATASETC", "dataMontlydata", "Check")
os.makedirs(CHECK_DIR, exist_ok=True)
CHECK_CSV = os.path.join(CHECK_DIR, "DataMontlyCheck.csv")
DATA_CSV = os.path.join(os.getcwd(), "DATASETC", "dataMontlydata", "dataMontly.csv")
DISPLAY_COLUMNS = ["Chủng loại", "Mã hàng", "Khách hàng", "Link", "Status"]


def open_config_monthly_window(root):
    window = tk.Toplevel(root)
    window.title("Config Monthly Data")
    window.geometry("900x420")
    window.configure(bg="#e8ecef")

    from ult.SendEmail.Guidle.config import load_monthly_config, save_monthly_config

    config = load_monthly_config()

    # ====== Địa chỉ file nguồn ======
    frame_tempt = tk.Frame(window, bg="#e8ecef")
    frame_tempt.pack(pady=15, fill="x", padx=30)
    tk.Label(frame_tempt, text="Chọn thư mục tạm:", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
    entry_tempt = tk.Entry(frame_tempt, width=40, font=("Helvetica", 12))
    entry_tempt.pack(side=tk.LEFT, padx=10)
    entry_tempt.insert(0, config.get("tempt_path", ""))
    def select_tempt_folder():
        folder = filedialog.askdirectory(title="Chọn thư mục tạm")
        if folder:
            entry_tempt.delete(0, tk.END)
            entry_tempt.insert(0, folder)
    tk.Button(frame_tempt, text="Chọn", command=select_tempt_folder,
            font=("Helvetica", 11, "bold"), bg="#3498db", fg="white", padx=10).pack(side=tk.LEFT)


    # ====== Địa chỉ file gốc ======
    frame_origin = tk.Frame(window, bg="#e8ecef")
    frame_origin.pack(pady=15, fill="x", padx=30)
    tk.Label(frame_origin, text="Địa chỉ file gốc:", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
    entry_origin = tk.Entry(frame_origin, width=40, font=("Helvetica", 12))
    entry_origin.pack(side=tk.LEFT, padx=10)
    entry_origin.insert(0, config.get("origin_path", ""))
    def select_origin_folder():
        folder = filedialog.askdirectory(title="Chọn thư mục gốc")
        if folder:
            entry_origin.delete(0, tk.END)
            entry_origin.insert(0, folder)
    tk.Button(frame_origin, text="Chọn", command=select_origin_folder,
              font=("Helvetica", 11, "bold"), bg="#3498db", fg="white", padx=10).pack(side=tk.LEFT)

    # ====== File Excel ======
    frame_excel = tk.Frame(window, bg="#e8ecef")
    frame_excel.pack(pady=15, fill="x", padx=30)
    tk.Label(frame_excel, text="File Excel:", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
    entry_excel = tk.Entry(frame_excel, width=32, font=("Helvetica", 12))
    entry_excel.pack(side=tk.LEFT, padx=10)
    entry_excel.insert(0, config.get("excel_path", ""))

    def select_excel_file():
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.csv"), ("All files", "*.*")]
        )
        if file_path:
            entry_excel.delete(0, tk.END)
            entry_excel.insert(0, file_path)

    tk.Button(frame_excel, text="Chọn", command=select_excel_file,
              font=("Helvetica", 11, "bold"), bg="#3498db", fg="white", padx=10).pack(side=tk.LEFT)

    def convert_excel_to_csv():
        excel_path = entry_excel.get()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel hợp lệ!")
            return
        try:
            xl = pd.ExcelFile(excel_path)
            sheet_names = xl.sheet_names

            select_window = tk.Toplevel(window)
            select_window.title("Chọn sheet và vùng dữ liệu")
            select_window.geometry("420x320")
            select_window.configure(bg="#e8ecef")

            tk.Label(select_window, text="Chọn sheet:", font=("Helvetica", 12), bg="#e8ecef").pack(pady=8)
            sheet_var = tk.StringVar(value=sheet_names[0])
            sheet_menu = ttk.Combobox(select_window, textvariable=sheet_var, values=sheet_names, state="readonly", font=("Helvetica", 12))
            sheet_menu.pack(pady=5)

            row_frame = tk.Frame(select_window, bg="#e8ecef")
            row_frame.pack(pady=5)
            tk.Label(row_frame, text="Dòng bắt đầu (từ 0):", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
            entry_row_start = tk.Entry(row_frame, width=6, font=("Helvetica", 12))
            entry_row_start.pack(side=tk.LEFT, padx=5)
            entry_row_start.insert(0, "0")

            tk.Label(row_frame, text="Dòng kết thúc (từ 0, để trống lấy hết):", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
            entry_row_end = tk.Entry(row_frame, width=6, font=("Helvetica", 12))
            entry_row_end.pack(side=tk.LEFT, padx=5)

            col_frame = tk.Frame(select_window, bg="#e8ecef")
            col_frame.pack(pady=5)
            tk.Label(col_frame, text="Cột bắt đầu (từ 0):", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
            entry_col_start = tk.Entry(col_frame, width=6, font=("Helvetica", 12))
            entry_col_start.pack(side=tk.LEFT, padx=5)
            entry_col_start.insert(0, "0")

            tk.Label(col_frame, text="Cột kết thúc (từ 0, để trống lấy hết):", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT)
            entry_col_end = tk.Entry(col_frame, width=6, font=("Helvetica", 12))
            entry_col_end.pack(side=tk.LEFT, padx=5)

            def do_convert():
                try:
                    sheet = sheet_var.get()
                    row_start = entry_row_start.get().strip()
                    row_end = entry_row_end.get().strip()
                    col_start = entry_col_start.get().strip()
                    col_end = entry_col_end.get().strip()

                    df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
                    row_start = int(row_start) if row_start else 0
                    row_end = int(row_end) if row_end else None
                    col_start = int(col_start) if col_start else 0
                    col_end = int(col_end) if col_end else None

                    df = df.iloc[row_start:row_end, col_start:col_end]

                    base_name = os.path.splitext(os.path.basename(excel_path))[0]
                    output_csv = os.path.join(os.path.dirname(excel_path), f"{base_name}_monthly_convert.csv")
                    df.to_csv(output_csv, index=False, header=False, encoding='utf-8-sig')

                    # Lưu lại đường dẫn vào config
                    config = load_monthly_config()
                    config["excel_path"] = excel_path
                    config["excel_csv"] = output_csv
                    save_monthly_config(config)
                    messagebox.showinfo("Thành công", f"Đã convert file Excel sang CSV:\n{output_csv}")
                    select_window.destroy()
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể convert file Excel: {e}")

            tk.Button(select_window, text="Convert", command=do_convert,
                      font=("Helvetica", 12, "bold"), bg="#f39c12", fg="white", padx=18).pack(pady=18)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file Excel: {e}")

    tk.Button(frame_excel, text="Convert to CSV", command=convert_excel_to_csv,
              font=("Helvetica", 11, "bold"), bg="#f39c12", fg="white", padx=10).pack(side=tk.LEFT, padx=8)

    # ====== Nút lưu và đóng ======
    frame_btn = tk.Frame(window, bg="#e8ecef")
    frame_btn.pack(pady=30)
    def save():
        config = {
        "tempt_path": entry_tempt.get(),
        "origin_path": entry_origin.get(),
        "excel_path": entry_excel.get()
    }
        save_monthly_config(config)
        messagebox.showinfo("Thành công", "Đã lưu cấu hình Monthly Data!")

    tk.Button(frame_btn, text="Lưu cấu hình", command=save, font=("Helvetica", 12, "bold"),
              bg="#27ae60", fg="white", padx=20, pady=6).pack(side=tk.LEFT, padx=20)
    tk.Button(frame_btn, text="Đóng", command=window.destroy, font=("Helvetica", 12, "bold"),
              bg="#e74c3c", fg="white", padx=20, pady=6).pack(side=tk.LEFT, padx=20)

def update_check_data():
    # Xóa file cũ nếu có
    if os.path.exists(CHECK_CSV):
        os.remove(CHECK_CSV)
    # Đọc dataMontly.csv
    if os.path.exists(DATA_CSV):
        df = pd.read_csv(DATA_CSV, encoding="utf-8-sig")
        # Thêm cột Link, Status nếu chưa có
        if "Link" not in df.columns:
            df["Link"] = ""
        if "Status" not in df.columns:
            df["Status"] = ""
        if "Link nguồn" not in df.columns:
            df["Link nguồn"] = ""
        df = df[["Chủng loại", "Mã hàng", "Khách hàng", "Link", "Status", "Link nguồn"]]
        df.to_csv(CHECK_CSV, index=False, encoding="utf-8-sig")
    else:
        df = pd.DataFrame(columns=DISPLAY_COLUMNS + ["Link", "Status", "Link nguồn"])
        df.to_csv(CHECK_CSV, index=False, encoding="utf-8-sig")

def load_check_data():
    if os.path.exists(CHECK_CSV):
        try:
            return pd.read_csv(CHECK_CSV, encoding="utf-8-sig")
        except Exception:
            return pd.DataFrame(columns=DISPLAY_COLUMNS)
    return pd.DataFrame(columns=DISPLAY_COLUMNS)

def save_check_data(df):
    df.to_csv(CHECK_CSV, index=False, encoding="utf-8-sig")
    
    

  
def open_gui_monthly_data(root, parent_window=None):
    window = tk.Toplevel(root)
    window.title("Gửi Monthly Data")
    window.geometry("1200x700")
    window.configure(bg="#e8ecef")
    window.lift()
    window.grab_set()
    window.focus_force()
    if parent_window:
        parent_window.withdraw()
        def on_close():
            parent_window.deiconify()
            window.destroy()
        window.protocol("WM_DELETE_WINDOW", on_close)

    
    # ==== Khu vực chọn file KJS và chọn tháng ====
    frame_top = tk.Frame(window, bg="#e8ecef")
    frame_top.pack(pady=10, fill="x")

    # Chọn file KJS
    tk.Label(frame_top, text="Chọn file KJS:", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT, padx=(10,0))
    entry_kjs = tk.Entry(frame_top, width=40, font=("Helvetica", 12))
    entry_kjs.pack(side=tk.LEFT, padx=5)
    def select_kjs_file():
        path = filedialog.askopenfilename(title="Chọn file KJS", filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All files", "*.*")])
        if path:
            entry_kjs.delete(0, tk.END)
            entry_kjs.insert(0, path)
    tk.Button(frame_top, text="Chọn", command=select_kjs_file, font=("Helvetica", 11, "bold"), bg="#3498db", fg="white", padx=10).pack(side=tk.LEFT, padx=5)

    # Chọn tháng
    tk.Label(frame_top, text="Chọn tháng:", font=("Helvetica", 12), bg="#e8ecef").pack(side=tk.LEFT, padx=(30,0))
    month_var = tk.StringVar()
    
    def pick_month():
        top = tk.Toplevel(window)
        top.title("Chọn tháng")
        top.geometry("300x350")
        top.configure(bg="#e8ecef")

        tk.Label(top, text="Tháng:", font=("Helvetica", 12), bg="#e8ecef").pack(pady=5)
        month_cb = ttk.Combobox(top, values=[f"{i:02d}" for i in range(1, 13)], width=5, font=("Helvetica", 12), state="readonly")
        month_cb.pack()
        month_cb.set(datetime.datetime.now().strftime("%m"))

        tk.Label(top, text="Năm:", font=("Helvetica", 12), bg="#e8ecef").pack(pady=5)
        year_cb = ttk.Combobox(top, values=[str(y) for y in range(datetime.datetime.now().year-3, datetime.datetime.now().year+4)], width=7, font=("Helvetica", 12), state="readonly")
        year_cb.pack()
        year_cb.set(datetime.datetime.now().strftime("%Y"))

        def set_month():
            month = month_cb.get()
            year = year_cb.get()
            month_var.set(f"{month}/{year}")
            # Lưu vào stateMontly
            stateMontly.MonthSelect = month
            stateMontly.YearsSelect = year
            top.destroy()
        tk.Button(top, text="Chọn", command=set_month, font=("Helvetica", 12, "bold"), bg="#27ae60", fg="white").pack(pady=10)

    entry_month = tk.Entry(frame_top, textvariable=month_var, width=10, font=("Helvetica", 12))
    entry_month.pack(side=tk.LEFT, padx=5)
    tk.Button(frame_top, text="Chọn tháng", command=pick_month, font=("Helvetica", 11, "bold"), bg="#3498db", fg="white", padx=10).pack(side=tk.LEFT, padx=5)

    # Nút Update dữ liệu
    def update_data():
        update_check_data()
        nonlocal full_df
        full_df = load_check_data()
        refresh_check_tree()

    # TreeView
    frame_table = tk.Frame(window, bg="#e8ecef")
    frame_table.pack(pady=10, fill="both", expand=True)
    tree = ttk.Treeview(frame_table, columns=DISPLAY_COLUMNS, show="headings", height=20)
    for col in DISPLAY_COLUMNS:
        tree.heading(col, text=col)
        tree.column(col, width=200, anchor="center")
    tree.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    # Load data
    df = load_check_data()
    full_df = df.copy()  # DataFrame gốc, luôn giữ nguyên
    def filter_check_df():
        filtered = full_df.copy()
        for col in DISPLAY_COLUMNS:
            val = filter_vars[col].get().strip()
            if val:
                filtered = filtered[filtered[col].astype(str).str.contains(val, case=False, na=False)]
        return filtered

    def refresh_check_tree(data=None):
        tree.delete(*tree.get_children())
        if data is None:
            data = full_df
        for _, row in data.iterrows():
            tree.insert(
                "",
                "end",
                values=(
                    row.get("Chủng loại", ""),
                    row.get("Mã hàng", ""),
                    row.get("Khách hàng", ""),
                    row.get("Link", ""),
                    row.get("Status", "")
                )
            )
    refresh_check_tree()
    
    def nen_file():
        # Lấy tháng/năm từ stateMontly
        month = stateMontly.MonthSelect or datetime.datetime.now().strftime("%m")
        year = stateMontly.YearsSelect or datetime.datetime.now().strftime("%Y")
        # Thư mục gốc cần nén
        from ult.SendEmail.Guidle.config import load_monthly_config
        config = load_monthly_config()
        tempt_dir = config.get("tempt_path", "")
        origin_dir = config.get("origin_path", "")
        base_dir = os.path.join(config.get("tempt_path", ""), year, month)
        if not os.path.exists(base_dir):
            messagebox.showwarning("Thiếu thư mục", f"Không tìm thấy thư mục: {base_dir}")
            return

        khach_hangs = ["CANON", "HP", "DENSO"]
        zipped = []
        for kh in khach_hangs:
            folder = os.path.join(base_dir, kh)
            if not os.path.exists(folder):
                continue
            zip_path = os.path.join(base_dir, f"{kh}_{month}.zip")
            try:
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root_, dirs, files in os.walk(folder):
                        for file in files:
                            file_path = os.path.join(root_, file)
                            arcname = os.path.relpath(file_path, folder)
                            zipf.write(file_path, arcname)
                zipped.append(zip_path)
            except Exception as e:
                messagebox.showerror("Lỗi", f"Lỗi nén {kh}: {e}")
        if zipped:
            messagebox.showinfo("Thành công", f"Đã nén: \n" + "\n".join(zipped))
        else:
            messagebox.showwarning("Không có dữ liệu", "Không tìm thấy thư mục khách hàng để nén!")
    
    def filter_check_df():
        filtered = full_df.copy()
        for col in DISPLAY_COLUMNS:
            val = filter_vars[col].get().strip()
            if val:
                filtered = filtered[filtered[col].astype(str).str.contains(val, case=False, na=False)]
        return filtered

    def refresh_check_tree(data=None):
        tree.delete(*tree.get_children())
        if data is None:
            data = full_df
        for _, row in data.iterrows():
            tree.insert(
                "",
                "end",
                values=(
                    row.get("Chủng loại", ""),
                    row.get("Mã hàng", ""),
                    row.get("Khách hàng", ""),
                    row.get("Link", ""),
                    row.get("Status", "")
                )
            )


    def confirm_data():
        kjs_path = entry_kjs.get().strip()
        if not kjs_path or not os.path.exists(kjs_path):
            messagebox.showwarning("Thiếu file", "Vui lòng chọn file KJS hợp lệ!")
            return

        try:
            # Đọc file KJS (Excel hoặc CSV)
            if kjs_path.lower().endswith(('.xlsx', '.xls')):
                kjs_df = pd.read_excel(kjs_path, dtype=str)
            else:
                kjs_df = pd.read_csv(kjs_path, dtype=str, encoding="utf-8-sig")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file KJS: {e}")
            return

        # Kiểm tra cột ITEM và PRODUCTION_ORDER_NO
        if "ITEM" not in kjs_df.columns or "PRODUCTION_ORDER_NO" not in kjs_df.columns:
            messagebox.showerror("Lỗi", "File KJS phải có cột 'ITEM' và 'PRODUCTION_ORDER_NO'!")
            return

        # Đọc DataMontlyCheck.csv
        df = load_check_data()
        found_count = 0
        filter_rows = []
        json_dict = {}

        for idx, row in df.iterrows():
            ma_hang = str(row["Mã hàng"])
            # Chủng loại: lấy 2 phần đầu, ghép lại không dấu "-"
            parts = ma_hang.split("-")
            if len(parts) >= 2:
                chung_loai = parts[0] + parts[1]
                ChungloaiMini = parts[1]
                tenmahangMini=parts[2]
            else:
                chung_loai = ma_hang
                ChungloaiMini=""
                tenmahangMini=""
            khach_hang = str(row["Khách hàng"])  # lấy từ dòng hiện tại

            # Lấy phần cuối sau dấu "-"
            if "-" in ma_hang:
                code = ma_hang.split("-")[-1].strip()
            else:
                code = ma_hang.strip()
            # Lọc các dòng KJS có ITEM chứa code và PRODUCTION_ORDER_NO > 6 ký tự
            matched = kjs_df[
                kjs_df["ITEM"].astype(str).str.contains(code, case=False, na=False) &
                (kjs_df["PRODUCTION_ORDER_NO"].astype(str).str.len() > 6)
            ]
            if not matched.empty:
                df.at[idx, "Status"] = "Xác nhận có dữ liệu KJS"
                found_count += 1
                for _, kjs_row in matched.iterrows():
                    filter_rows.append(kjs_row)
                    item_key = str(kjs_row["ITEM"])
                    item_info = {
                        "ChungLoai": chung_loai,
                        "ChungloaiMini": ChungloaiMini,
                        "TenmahangMini": tenmahangMini,
                        "KhachHang": khach_hang,
                        "LOT_NO": kjs_row.get("LOT_NO", ""),
                        "CUSTOMER": kjs_row.get("CUSTOMER", ""),
                        "ACCEPT_QTY": kjs_row.get("ACCEPT_QTY", ""),
                        "PRODUCTION_ORDER_NO": kjs_row.get("PRODUCTION_ORDER_NO", "")
                    }
                    if item_key not in json_dict:
                        json_dict[item_key] = []
                    json_dict[item_key].append(item_info)
            else:
                df.at[idx, "Status"] = "Không có dữ liệu trong KJS"

        # Lưu lại DataMontlyCheck.csv
        save_check_data(df)
        nonlocal full_df
        full_df = load_check_data()
        refresh_check_tree()

        # Lưu DataMontlyFilter.csv
        if filter_rows:
            filter_df = pd.DataFrame(filter_rows)
            filter_csv_path = os.path.join(CHECK_DIR, "DataMontlyFilter.csv")
            filter_df.to_csv(filter_csv_path, index=False, encoding="utf-8-sig")
            # Lưu JSON
            import json
            filter_json_path = os.path.join(CHECK_DIR, "DataMontlyFilter.json")
            with open(filter_json_path, "w", encoding="utf-8") as f:
                json.dump(json_dict, f, ensure_ascii=False, indent=2)
        else:
            filter_csv_path = ""
            filter_json_path = ""

        messagebox.showinfo("Kết quả", f"Đã xác nhận xong!\nCó {found_count} mã hàng tìm thấy dữ liệu KJS.\n"
                                    f"Đã xuất file DataMontlyFilter.csv{' và DataMontlyFilter.json' if filter_rows else ''}.")
    
    def open_file_from_tree(event):
        selected = tree.selection()
        if not selected:
            return
        item = selected[0]
        values = tree.item(item, "values")
        excel_path = values[3]  # Cột "Link"
        if excel_path and os.path.exists(excel_path):
            os.startfile(excel_path)
        else:
            messagebox.showerror("Lỗi", f"Không tìm thấy file: {excel_path}")
    tree.bind("<Double-1>", open_file_from_tree)
    
    filter_vars = {col: tk.StringVar() for col in DISPLAY_COLUMNS}
    def apply_filters():
        filtered = df.copy()
        for col in DISPLAY_COLUMNS:
            val = filter_vars[col].get().strip()
            if val:
                filtered = filtered[filtered[col].astype(str).str.contains(val, case=False, na=False)]        
        return filtered

    def show_filter(col):
        filter_win = tk.Toplevel(window)
        filter_win.title(f"Lọc theo {col}")
        filter_win.geometry("400x180")
        filter_win.configure(bg="#e8ecef")
        filter_win.lift()
        filter_win.grab_set()
        tk.Label(filter_win, text=f"Nhập giá trị lọc cho {col}:", font=("Helvetica", 12), bg="#e8ecef").pack(pady=15)
        entry = tk.Entry(filter_win, width=35, font=("Helvetica", 12))
        entry.pack(pady=10)
        entry.insert(0, filter_vars[col].get())
        def apply_filter():
            filter_vars[col].set(entry.get().strip())
            filtered_data = filter_check_df()
            refresh_check_tree(filtered_data)
            filter_win.destroy()
        tk.Button(filter_win, text="Lọc", command=apply_filter, font=("Helvetica", 12, "bold"),
                bg="#3498db", fg="white", padx=20, pady=8).pack(pady=15)

    for col in DISPLAY_COLUMNS:
        tree.heading(col, text=col, command=lambda c=col: show_filter(c))
    
    
       
    def edit_content():
        from ult.SendEmail.Guidle.config import load_monthly_config
        config = load_monthly_config()
        tempt_dir = config.get("tempt_path", "")
        origin_dir = config.get("origin_path", "")

        import ult.FileMontlyData.Guidle.stateMontly as state_monthly
        month = state_monthly.MonthSelect or datetime.datetime.now().strftime("%m")
        year = state_monthly.YearsSelect or datetime.datetime.now().strftime("%Y")

        # Đọc DataMontlyCheck.csv
        df = load_check_data()
        copied_files = []
        for idx, row in df.iterrows():
            if str(row["Status"]) != "Xác nhận có dữ liệu KJS":
                continue
            chungloaiMini = ""
            tenmahangMini = ""
            parts = str(row["Mã hàng"]).split("-")
            if len(parts) >= 3:
                chungloaiMini = parts[1]
                tenmahangMini = parts[2]
            khach_hang = str(row["Khách hàng"])
            src_folder = os.path.join(origin_dir, f"Hang {chungloaiMini}", f"Ma Hang {tenmahangMini}", year)
            if not os.path.exists(src_folder):
                continue
            pattern = f"{tenmahangMini}-{year}.{month}"
            found_link = ""
            for fname in os.listdir(src_folder):
                if pattern in fname and fname.endswith(('.xls', '.xlsx')):
                    src_file = os.path.join(src_folder, fname)
                    dest_folder = os.path.join(tempt_dir, year, month, khach_hang)
                    os.makedirs(dest_folder, exist_ok=True)
                    dest_file = os.path.join(dest_folder, fname)
                    if not os.path.exists(dest_file):
                        try:
                            import shutil
                            shutil.copy2(src_file, dest_file)
                            copied_files.append(dest_file)
                            df.at[idx, "Link nguồn"] = src_file  # Lưu lại đường dẫn nguồn
                        except Exception as e:
                            print(f"Lỗi copy {src_file}: {e}")
                    found_link = dest_file  # Lưu đường dẫn file vừa copy
                    break  # Nếu chỉ lấy 1 file đầu tiên khớp
            # Cập nhật cột Link cho dòng hiện tại
            df.at[idx, "Link"] = found_link
        save_check_data(df)
        nonlocal full_df
        full_df = load_check_data()
        refresh_check_tree()    
        messagebox.showinfo("Kết quả", f"Đã copy {len(copied_files)} file vào thư mục tạm.")

        # Sau khi copy xong, tiếp tục các bước chỉnh sửa dữ liệu như hiện tại
        # ... (phần xử lý chỉnh sửa file .xls/.xlsx giữ nguyên như bạn đã viết ở dưới) ...
        
        df = load_check_data()
        filter_json_path = os.path.join(CHECK_DIR, "DataMontlyFilter.json")
        with open(filter_json_path, "r", encoding="utf-8") as f:
            filter_data = json.load(f)
            
        total_files = 0
        completed_files = 0
        empty_files = 0
        error_files = 0

        for idx, row in df.iterrows():
            if str(row["Status"]) != "Xác nhận có dữ liệu KJS":
                continue
            ma_hang = str(row["Mã hàng"])
            khach_hang = str(row["Khách hàng"])
            parts = ma_hang.split("-")
            if len(parts) >= 3:
                chungloaiMini = parts[1]
                tenmahangMini = parts[2]
            else:
                continue

            dest_folder = os.path.join(tempt_dir, year, month, khach_hang)
            pattern = f"{tenmahangMini}-{year}.{month}"
            excel_files = [f for f in os.listdir(dest_folder) if pattern in f and f.endswith(('.xls', '.xlsx'))]
            if not excel_files:
                df.at[idx, "Status"] = "Không tìm thấy file"
                error_files += 1
                continue
            for excel_file in excel_files:
                file_path = os.path.join(dest_folder, excel_file)
                total_files += 1
                try:
                    found_lot = False
                    checked_lot = set()
                    code = ma_hang.split("-")[-1].strip()
                    lot_list = []
                    for item_key, lots in filter_data.items():
                        if code in item_key:
                            for lot in lots:
                                lot_list.append(lot["LOT_NO"].replace("-", ""))
                    # Xử lý file .xlsx
                    if excel_file.lower().endswith('.xlsx'):
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        ws = wb.active
                        cell_C8 = ws["C8"].value
                        if cell_C8 != ma_hang:
                            df.at[idx, "Status"] = "File lỗi dữ liệu"
                            error_files += 1
                            continue
                        for id_lot, start_row in enumerate(range(25, 1000, 39)):
                            cell_E = ws[f"E{start_row}"].value
                            if not cell_E:
                                df.at[idx, "Status"] = "Dữ liệu bị trống"
                                empty_files += 1
                                break
                            lot_no_excel = str(cell_E).replace("-", "")
                            if lot_no_excel in checked_lot:
                                continue
                            checked_lot.add(lot_no_excel)
                            if lot_no_excel in lot_list:
                                error_cells = []
                                for j in range(start_row+32, start_row-1, -1):
                                    val = ws[f"I{j}"].value
                                    if val in ["#DIV/0!"]:
                                        error_cells.append(j)
                                for j in error_cells:
                                    for col in ["I", "J", "K", "L", "M"]:
                                        ws[f"{col}{j}"].value = None
                                found_lot = True
                                if id_lot == 0 and len(lot_list) == 1:
                                    for r in range(start_row+39, 1001):
                                        for col in ["A", "B", "C", "D", "E"]:
                                            ws[f"{col}{r}"].value = None
                                break
                        if found_lot:
                            order_no = ws["P1"].value
                            if order_no and "ORDER No:" in str(order_no):
                                pass
                            for j in range(25, 6, -1):
                                val = ws[f"X{j}"].value
                                if val in ["#DIV/0!"]:
                                    for col in range(ord("P"), ord("A")+24):
                                        ws[f"{chr(col)}{j}"].value = None
                            wb.save(file_path)
                            df.at[idx, "Status"] = "Hoàn thành chỉnh sửa dữ liệu"
                            completed_files += 1
                        elif not found_lot:
                            df.at[idx, "Status"] = "File lỗi dữ liệu"
                            error_files += 1

                    # Xử lý file .xls
                    elif excel_file.lower().endswith('.xls'):
                        app = xw.App(visible=False)
                        wb = app.books.open(file_path)
                        ws = wb.sheets[0]
                        
                        khach_hang = str(row["Khách hàng"]).strip().upper()
                        json_key = f"{parts[0]}{parts[1]}{parts[2]}"
                        lots = filter_data.get(json_key, [])
                        order_no_list = []
                        found_lot = False
                        part_ranges_to_delete = []  # Lưu vùng cần xóa sau khi duyệt
                        
                        if khach_hang == "CANON":
                            cell_C8 = ws.range("C8").value
                            if cell_C8 != ma_hang:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1
                                wb.close()
                                app.quit()
                                continue
                            order_no_list = []
                            found_lot = False
                            
                            # Xét từng phần (E25, E64, E103, ...)
                            for id_lot, start_row in enumerate(range(25, 1000, 39)):
                                cell_E = ws.range(f"E{start_row}").value
                                if not cell_E:
                                    part_ranges_to_delete.append((start_row, start_row+38))
                                    continue
                                lot_no_excel = str(cell_E).replace("-", "")
                                lot_info = next((lot for lot in lots if lot["LOT_NO"].replace("-", "") == lot_no_excel), None)
                                if lot_info:
                                    # ws.range(f"H{start_row}").value = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    ws.range(f"J{start_row}").value = lot_info.get("ACCEPT_QTY", "")
                                    # ws.range(f"L{start_row}").value = lot_info.get("CUSTOMER", "")
                                    po_no = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    if po_no and po_no not in order_no_list:
                                        order_no_list.append(po_no)
                                    found_lot = True

                            # Xét các dòng W7 -> W32 (của toàn sheet)
                            rows_to_delete = []
                            for j in range(7, 32):  # W7 đến W32 (bao gồm cả 31)
                                cell = ws.range(f"W{j}")
                                cell_W = cell.value
                                cell_formula = cell.formula
                                
                                if isinstance(cell_W, str) and cell_W.startswith("#"):
                                    rows_to_delete.append(j)
                                elif cell_W is None:
                                    if isinstance(cell_formula, str) and cell_formula.startswith("="):
                                        # Công thức bị lỗi (giá trị None nhưng là công thức)
                                        rows_to_delete.append(j)

                            # Xóa các dòng W7-W31 theo range P-AX, xóa từ dòng lớn đến nhỏ
                            for j in sorted(rows_to_delete, reverse=True):
                                ws.range(f"P{j}:AX{j}").delete(shift="up")

                            # Xóa các vùng phần không có dữ liệu (sau khi duyệt xong)
                            for start, end in sorted(part_ranges_to_delete, reverse=True):
                                ws.range(f"A{start}:M{end}").delete(shift="up") 

                            # Ghi vào ô P1 chỉ các mã PO khác nhau
                            order_no_text = "ORDER No:"
                            if order_no_list:
                                ws.range("P1").value = f"{order_no_text} {', '.join(order_no_list)}"

                            wb.save()
                            wb.close()
                            app.quit()
                            if found_lot:
                                df.at[idx, "Status"] = "Hoàn thành chỉnh sửa dữ liệu"
                                completed_files += 1
                            else:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1

                        elif khach_hang == "HP":
                            cell_C14 = ws.range("D14").value
                            if cell_C14 != ma_hang:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1
                                wb.close()
                                app.quit()
                                continue
                            order_no_list = []
                            found_lot = False
                            part_ranges_to_delete = []
                            # Xét từng phần (E31, E70, E109, ...)
                            for id_lot, start_row in enumerate(range(31, 3930, 39)):
                                cell_E = ws.range(f"E{start_row}").value
                                if not cell_E:
                                    part_ranges_to_delete.append((start_row, start_row+38))
                                    continue
                                lot_no_excel = str(cell_E).replace("-", "")
                                lot_info = next((lot for lot in lots if lot["LOT_NO"].replace("-", "") == lot_no_excel), None)
                                if lot_info:
                                    # ws.range(f"I{start_row}").value = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    ws.range(f"L{start_row}").value = lot_info.get("ACCEPT_QTY", "")
                                    # ws.range(f"O{start_row}").value = lot_info.get("CUSTOMER", "")
                                    # ws.range(f"R{start_row}").value = cell_E  # Giá trị hiện tại của E
                                    po_no = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    if po_no and po_no not in order_no_list:
                                        order_no_list.append(po_no)
                                    found_lot = True
                            # Xét các dòng W7 -> W32 (của toàn sheet)
                            rows_to_delete = []
                            for j in range(13, 38):  # W7 đến W32 (bao gồm cả 31)
                                cell = ws.range(f"AA{j}")
                                cell_W = ws.range(f"AA{j}").value
                                cell_formula = cell.formula
                                if isinstance(cell_W, str) and cell_W.startswith("#"):
                                    rows_to_delete.append(j)
                                elif cell_W is None:
                                    if isinstance(cell_formula, str) and cell_formula.startswith("="):
                                        # Công thức bị lỗi (giá trị None nhưng là công thức)
                                        rows_to_delete.append(j)

                            # Xóa các dòng W7-W31 theo range P-AX, xóa từ dòng lớn đến nhỏ
                            for j in sorted(rows_to_delete, reverse=True):
                                ws.range(f"R{j}:AN{j}").delete(shift="up")

                            # Xóa các vùng phần không có dữ liệu (sau khi duyệt xong)
                            for start, end in sorted(part_ranges_to_delete, reverse=True):
                                ws.range(f"A{start}:T{end}").delete(shift="up") 
                            # Ghi vào ô P1
                            order_no_text = "ORDER No:"
                            if order_no_list:
                                ws.range("U3").value = f"{order_no_text} {', '.join(order_no_list)}"
                            wb.save()
                            wb.close()
                            app.quit()
                            if found_lot:
                                df.at[idx, "Status"] = "Hoàn thành chỉnh sửa dữ liệu (HP)"
                                completed_files += 1
                            else:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1

                        elif khach_hang == "DENSO":
                            cell_C24 = ws.range("C8").value
                            if cell_C24 != ma_hang:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1
                                wb.close()
                                app.quit()
                                continue
                            order_no_list = []
                            found_lot = False
                            part_ranges_to_delete = []
                            # Xét từng phần (E24, E43, E62, ...)
                            for id_lot, start_row in enumerate(range(24, 593, 19)):
                                cell_E = ws.range(f"E{start_row}").value
                                if not cell_E:
                                    part_ranges_to_delete.append((start_row, start_row+18))
                                    continue
                                lot_no_excel = str(cell_E).replace("-", "")
                                lot_info = next((lot for lot in lots if lot["LOT_NO"].replace("-", "") == lot_no_excel), None)
                                if lot_info:
                                    ws.range(f"H{start_row}").value = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    ws.range(f"J{start_row}").value = lot_info.get("ACCEPT_QTY", "")
                                    ws.range(f"L{start_row}").value = lot_info.get("CUSTOMER", "")
                                    po_no = lot_info.get("PRODUCTION_ORDER_NO", "")
                                    if po_no and po_no not in order_no_list:
                                        order_no_list.append(po_no)
                                    found_lot = True
                            # Xét các dòng W7 -> W32 (của toàn sheet)
                            rows_to_delete = []
                            for j in range(7, 32):  # W7 đến W32 (bao gồm cả 31)
                                cell = ws.range(f"AC{j}")
                                cell_W = ws.range(f"AC{j}").value
                                cell_formula = cell.formula

                                if isinstance(cell_W, str) and cell_W.startswith("#"):
                                    rows_to_delete.append(j)
                                elif cell_W is None:
                                    if isinstance(cell_formula, str) and cell_formula.startswith("="):
                                        # Công thức bị lỗi (giá trị None nhưng là công thức)
                                        rows_to_delete.append(j)

                            # Xóa các dòng W7-W31 theo range P-AX, xóa từ dòng lớn đến nhỏ
                            for j in sorted(rows_to_delete, reverse=True):
                                ws.range(f"N{j}:AD{j}").delete(shift="up")

                            # Xóa các vùng phần không có dữ liệu (sau khi duyệt xong)
                            for start, end in sorted(part_ranges_to_delete, reverse=True):
                                ws.range(f"A{start}:L{end}").delete(shift="up") 
                            # Ghi vào ô N1
                            order_no_text = "ORDER No:"
                            if order_no_list:
                                ws.range("N1").value = f"{order_no_text} {', '.join(order_no_list)}"
                            # Lưu file
                            wb.save()
                            wb.close()
                            app.quit()
                            if found_lot:
                                df.at[idx, "Status"] = "Hoàn thành chỉnh sửa dữ liệu (DENSO)"
                                completed_files += 1
                            else:
                                df.at[idx, "Status"] = "File lỗi dữ liệu"
                                error_files += 1

                        else:
                            wb.close()
                            app.quit()
                            df.at[idx, "Status"] = "Khách hàng không hỗ trợ"
                            error_files += 1
                except Exception as e:
                    df.at[idx, "Status"] = "File lỗi dữ liệu"
                    error_files += 1
                    print(f"Lỗi xử lý file {file_path}: {e}")

        save_check_data(df)
        full_df = load_check_data()
        refresh_check_tree()
        
        window = tk._get_default_root()
        import time
        window.lift()
        window.focus_force()
        window.update()
        time.sleep(0.2)  # Cho hệ điều hành kịp chuyển focus  # Đảm bảo cửa sổ Monthly Data lên trên cùng
        # Hiện messagebox và KHÔNG gọi lại window.lift() sau đó!
        messagebox.showinfo(
            "Kết quả",
            f"Đã xử lý {total_files} file.\n"
            f"Hoàn thành: {completed_files}\n"
            f"Trống: {empty_files}\n"
            f"Lỗi: {error_files}",
            parent=window
        )
                
    # Nút chức năng
    frame_btn = tk.Frame(window, bg="#e8ecef")
    frame_btn.pack(pady=10)
    
    tk.Button(frame_btn, text="Xác nhận dữ liệu", font=("Helvetica", 12, "bold"),
            bg="#27ae60", fg="white", padx=18, pady=6,
            command=confirm_data).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_btn, text="Chỉnh sửa nội dung", font=("Helvetica", 12, "bold"),
          bg="#3498db", fg="white", padx=18, pady=6,
          command=edit_content).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_btn, text="Nén file", font=("Helvetica", 12, "bold"),
          bg="#f39c12", fg="white", padx=18, pady=6,
          command=nen_file).pack(side=tk.LEFT, padx=10)
    # Thêm ở đây
    tk.Button(frame_btn, text="Di chuyển File", font=("Helvetica", 12, "bold"),
            bg="#f39c12", fg="white", padx=18, pady=6,
            command=move_files_back).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_btn, text="Reset", font=("Helvetica", 12, "bold"),
              bg="#e74c3c", fg="white", padx=18, pady=6,
              command=refresh_check_tree).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_btn, text="Update dữ liệu", font=("Helvetica", 12, "bold"),
              bg="#f39c12", fg="white", padx=18, pady=6,
              command=update_data).pack(side=tk.LEFT, padx=10)
    tk.Button(window, text="Đóng", command=window.destroy, font=("Helvetica", 12, "bold"),
              bg="#8e44ad", fg="white", padx=20, pady=8).pack(pady=12)
    tk.Button(window, text="Quay lại", command=on_close, font=("Helvetica", 12, "bold"),
            bg="#e74c3c", fg="white", padx=20, pady=8).pack(pady=12)
def move_files_back():
    df = load_check_data()
    moved_count = 0
    error_count = 0
    for idx, row in df.iterrows():
        link = str(row.get("Link", "")).strip()
        link_src = str(row.get("Link nguồn", "")).strip()
        if link and link_src and os.path.exists(link):
            try:
                # Đảm bảo thư mục nguồn tồn tại
                os.makedirs(os.path.dirname(link_src), exist_ok=True)
                import shutil
                shutil.copy2(link, link_src)
                moved_count += 1
            except Exception as e:
                print(f"Lỗi copy ngược {link} -> {link_src}: {e}")
                error_count += 1
    messagebox.showinfo("Kết quả", f"Đã di chuyển {moved_count} file về nguồn.\nLỗi: {error_count}")