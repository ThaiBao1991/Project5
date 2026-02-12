import os
import shutil
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pyautogui
import time
import threading
import subprocess
import pygetwindow as gw
import pyperclip  # Thư viện để copy clipboard

class ExcelRPACopier:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel RPA Copier - Enhanced")
        self.root.geometry("700x700")
        
        # Biến lưu đường dẫn
        self.source_file_path = tk.StringVar()
        self.destination_folder = tk.StringVar()
        
        # Biến cho nội dung nhập và ô Excel
        self.cell_content = tk.StringVar(value="Bảo đẹp trai")
        self.excel_cell = tk.StringVar(value="A20")
        
        # Biến cho delay
        self.open_wait_time = tk.DoubleVar(value=5.0)
        self.step_delay = tk.DoubleVar(value=1.0)
        
        # Biến cho hiển thị từng bước
        self.show_steps = tk.BooleanVar(value=True)
        
        # Biến lưu file mới đã tạo
        self.new_file_path = None
        
        # Biến cho phương pháp nhập tiếng Việt
        self.vietnamese_method = tk.StringVar(value="paste")  # paste hoặc type
        
        self.setup_ui()
        
    def setup_ui(self):
        # Tiêu đề
        title_label = tk.Label(self.root, text="Excel RPA Copier - Enhanced", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # Frame chọn file nguồn
        source_frame = ttk.LabelFrame(self.root, text="1. Chọn file Excel nguồn", padding=10)
        source_frame.pack(fill="x", padx=20, pady=5)
        
        ttk.Label(source_frame, text="Đường dẫn file nguồn:").grid(row=0, column=0, sticky="w")
        
        self.source_entry = ttk.Entry(source_frame, textvariable=self.source_file_path, width=50)
        self.source_entry.grid(row=1, column=0, padx=(0, 10), sticky="ew")
        
        ttk.Button(source_frame, text="Browse", 
                  command=self.browse_source_file).grid(row=1, column=1)
        
        # Frame chọn thư mục đích
        dest_frame = ttk.LabelFrame(self.root, text="2. Chọn thư mục đích", padding=10)
        dest_frame.pack(fill="x", padx=20, pady=5)
        
        ttk.Label(dest_frame, text="Thư mục đích:").grid(row=0, column=0, sticky="w")
        
        self.dest_entry = ttk.Entry(dest_frame, textvariable=self.destination_folder, width=50)
        self.dest_entry.grid(row=1, column=0, padx=(0, 10), sticky="ew")
        
        ttk.Button(dest_frame, text="Browse", 
                  command=self.browse_destination_folder).grid(row=1, column=1)
        
        # Frame nhập nội dung và ô Excel
        content_frame = ttk.LabelFrame(self.root, text="3. Cấu hình nội dung nhập", padding=10)
        content_frame.pack(fill="x", padx=20, pady=5)
        
        # Nội dung cần nhập
        ttk.Label(content_frame, text="Nội dung cần nhập vào Excel:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.content_entry = ttk.Entry(content_frame, textvariable=self.cell_content, width=30)
        self.content_entry.grid(row=0, column=1, padx=(0, 20))
        
        # Ô Excel
        ttk.Label(content_frame, text="Ô Excel (ví dụ: A20):").grid(row=0, column=2, sticky="w", padx=(0, 10))
        self.cell_entry = ttk.Entry(content_frame, textvariable=self.excel_cell, width=10)
        self.cell_entry.grid(row=0, column=3)
        
        # Frame cấu hình thời gian chờ
        time_frame = ttk.LabelFrame(self.root, text="4. Cấu hình thời gian chờ", padding=10)
        time_frame.pack(fill="x", padx=20, pady=5)
        
        # Thời gian chờ mở file Excel
        ttk.Label(time_frame, text="Thời gian chờ mở file (giây):").grid(row=0, column=0, sticky="w", padx=(0, 10))
        open_wait_spinbox = ttk.Spinbox(time_frame, from_=3.0, to=15.0, increment=1.0, 
                                       textvariable=self.open_wait_time, width=8)
        open_wait_spinbox.grid(row=0, column=1, padx=(0, 20))
        
        # Delay giữa các bước
        ttk.Label(time_frame, text="Delay giữa các bước (giây):").grid(row=0, column=2, sticky="w", padx=(0, 10))
        step_delay_spinbox = ttk.Spinbox(time_frame, from_=0.5, to=3.0, increment=0.1, 
                                        textvariable=self.step_delay, width=8)
        step_delay_spinbox.grid(row=0, column=3, padx=(0, 20))
        
        # Frame cấu hình tiếng Việt
        vietnamese_frame = ttk.LabelFrame(self.root, text="5. Cấu hình nhập tiếng Việt", padding=10)
        vietnamese_frame.pack(fill="x", padx=20, pady=5)
        
        ttk.Label(vietnamese_frame, text="Phương pháp nhập tiếng Việt:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        # Radio buttons cho phương pháp nhập
        paste_radio = ttk.Radiobutton(vietnamese_frame, text="Paste (Ctrl+V) - Khuyến nghị", 
                                     variable=self.vietnamese_method, value="paste")
        paste_radio.grid(row=0, column=1, padx=(0, 20))
        
        type_radio = ttk.Radiobutton(vietnamese_frame, text="Type (có thể lỗi font)", 
                                    variable=self.vietnamese_method, value="type")
        type_radio.grid(row=0, column=2)
        
        # Frame cấu hình hiển thị
        display_frame = ttk.LabelFrame(self.root, text="6. Cấu hình hiển thị", padding=10)
        display_frame.pack(fill="x", padx=20, pady=5)
        
        # Checkbox hiển thị từng bước
        self.show_steps_check = ttk.Checkbutton(display_frame, text="Hiển thị popup từng bước khi chạy", 
                                               variable=self.show_steps)
        self.show_steps_check.grid(row=0, column=0, padx=(0, 20))
        
        # Frame hiển thị tên file mới
        info_frame = ttk.LabelFrame(self.root, text="Thông tin file sẽ được tạo", padding=10)
        info_frame.pack(fill="x", padx=20, pady=5)
        
        # Grid layout cho info frame
        ttk.Label(info_frame, text="Folder mới sẽ được tạo:").grid(row=0, column=0, sticky="w")
        self.new_folder_label = tk.Label(info_frame, text="Test_dd-mm-yyyy", 
                                        font=("Arial", 10, "bold"), fg="green")
        self.new_folder_label.grid(row=0, column=1, sticky="w", padx=(10, 20))
        
        ttk.Label(info_frame, text="File Excel mới sẽ được tạo:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.new_file_label = tk.Label(info_frame, text="Test_dd-mm-yyyy.xlsx", 
                                      font=("Arial", 10, "bold"), fg="blue")
        self.new_file_label.grid(row=1, column=1, sticky="w", padx=(10, 0), pady=(5, 0))
        
        ttk.Label(info_frame, text="Nội dung sẽ nhập:").grid(row=0, column=2, sticky="w", padx=(20, 10))
        self.content_preview_label = tk.Label(info_frame, text="Bảo đẹp trai", 
                                             font=("Arial", 10, "bold"), fg="purple")
        self.content_preview_label.grid(row=0, column=3, sticky="w", padx=(10, 0))
        
        ttk.Label(info_frame, text="Vào ô:").grid(row=1, column=2, sticky="w", padx=(20, 10), pady=(5, 0))
        self.cell_preview_label = tk.Label(info_frame, text="A20", 
                                          font=("Arial", 10, "bold"), fg="orange")
        self.cell_preview_label.grid(row=1, column=3, sticky="w", padx=(10, 0), pady=(5, 0))
        
        # Preview tiếng Việt
        ttk.Label(info_frame, text="Preview:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        self.vietnamese_preview = tk.Label(info_frame, text="Bảo đẹp trai", 
                                          font=("Arial", 12, "bold"), 
                                          bg="lightyellow", padx=10, pady=5,
                                          borderwidth=2, relief="solid")
        self.vietnamese_preview.grid(row=2, column=1, columnspan=3, sticky="ew", pady=(10, 0), padx=(10, 0))
        
        # Frame các nút điều khiển
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=15)
        
        # Nút Test Vietnamese Input
        ttk.Button(button_frame, text="Test Tiếng Việt", 
                  command=self.test_vietnamese_input).pack(side="left", padx=5)
        
        # Nút Test Keyboard
        ttk.Button(button_frame, text="Test Keyboard", 
                  command=self.test_keyboard_actions).pack(side="left", padx=5)
        
        # Nút Create Test Folder
        ttk.Button(button_frame, text="Tạo Folder Test", 
                  command=self.create_test_folder).pack(side="left", padx=5)
        
        # Nút chỉ copy file
        ttk.Button(button_frame, text="Chỉ Copy File", 
                  command=self.copy_file_only).pack(side="left", padx=5)
        
        # Nút Start RPA
        self.start_button = ttk.Button(button_frame, text="START RPA", 
                                      command=self.start_rpa_process,
                                      style="Accent.TButton")
        self.start_button.pack(side="left", padx=5)
        
        # Nút Mở File Excel
        ttk.Button(button_frame, text="Mở File Excel", 
                  command=self.open_excel_file).pack(side="left", padx=5)
        
        # Frame hiển thị log
        log_frame = ttk.LabelFrame(self.root, text="Log hoạt động", padding=10)
        log_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Tạo Text widget cho log với scrollbar
        scrollbar = tk.Scrollbar(log_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.log_text = tk.Text(log_frame, height=12, yscrollcommand=scrollbar.set,
                               wrap="word", font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True)
        scrollbar.config(command=self.log_text.yview)
        
        # Thêm tag để highlight
        self.log_text.tag_config("step", foreground="blue", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("success", foreground="green", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("warning", foreground="orange", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("error", foreground="red", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("highlight", foreground="purple", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("vietnamese", foreground="darkgreen", font=("Consolas", 9, "bold"))
        
        # Định dạng nút START
        style = ttk.Style()
        style.configure("Accent.TButton", font=("Arial", 10, "bold"))
        
        # Cập nhật thông tin preview
        self.update_previews()
        
        # Bind sự kiện thay đổi nội dung
        self.cell_content.trace_add("write", lambda *args: self.update_previews())
        self.excel_cell.trace_add("write", lambda *args: self.update_previews())
        
    def browse_source_file(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel nguồn",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.source_file_path.set(file_path)
            self.log_message(f"Đã chọn file nguồn: {os.path.basename(file_path)}", "success")
            
    def browse_destination_folder(self):
        folder_path = filedialog.askdirectory(title="Chọn thư mục đích")
        if folder_path:
            self.destination_folder.set(folder_path)
            self.log_message(f"Đã chọn thư mục đích: {folder_path}", "success")
            
    def update_previews(self):
        """Cập nhật thông tin preview"""
        current_date = datetime.now().strftime("%d-%m-%Y")
        new_folder = f"Test_{current_date}"
        new_file = f"Test_{current_date}.xlsx"
        
        self.new_folder_label.config(text=new_folder)
        self.new_file_label.config(text=new_file)
        
        content = self.cell_content.get()
        self.content_preview_label.config(text=content)
        self.cell_preview_label.config(text=self.excel_cell.get())
        self.vietnamese_preview.config(text=content)
        
        # Kiểm tra nếu có tiếng Việt
        vietnamese_chars = "áàảãạăắằẳẵặâấầẩẫậđéèẻẽẹêếềểễệíìỉĩịóòỏõọôốồổỗộơớờởỡợúùủũụưứừửữựýỳỷỹỵ"
        has_vietnamese = any(char in content for char in vietnamese_chars)
        
        if has_vietnamese:
            self.vietnamese_preview.config(fg="red", font=("Arial", 12, "bold"))
        else:
            self.vietnamese_preview.config(fg="black", font=("Arial", 12, "bold"))
        
    def log_message(self, message, tag=None):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.log_text.insert("end", log_entry, tag)
        self.log_text.see("end")
        self.root.update()
        
    def create_test_folder(self):
        """Tạo folder Test với file Excel mẫu"""
        try:
            test_folder = os.path.join(os.getcwd(), "Test")
            os.makedirs(test_folder, exist_ok=True)
            
            excel_path = os.path.join(test_folder, "Test.xlsx")
            
            data = {
                'STT': [1, 2, 3, 4, 5],
                'Tên': ['Nguyễn Văn A', 'Trần Thị B', 'Lê Văn C', 'Phạm Thị D', 'Hoàng Văn E'],
                'Tuổi': [25, 30, 28, 35, 27],
                'Phòng Ban': ['IT', 'HR', 'Sales', 'Marketing', 'IT']
            }
            df = pd.DataFrame(data)
            df.to_excel(excel_path, index=False)
            
            # Thêm dữ liệu tiếng Việt
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font
                
                wb = load_workbook(excel_path)
                ws = wb.active
                
                ws['A1'] = "DANH SÁCH NHÂN VIÊN"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:D1')
                
                # Thêm dữ liệu tiếng Việt vào các ô test
                ws['A20'] = "Ô này sẽ bị ghi đè"
                ws['A21'] = "Tiếng Việt có dấu: á à ả ã ạ"
                ws['B20'] = "Chào mừng bạn đến với RPA"
                ws['C20'] = "Hà Nội, Việt Nam"
                
                wb.save(excel_path)
            except Exception as e:
                self.log_message(f"Không thể thêm định dạng: {str(e)}", "warning")
            
            self.log_message(f"✓ Đã tạo folder Test tại: {test_folder}", "success")
            self.log_message(f"✓ Đã tạo file Excel mẫu: Test.xlsx", "success")
            self.log_message(f"✓ File có sẵn dữ liệu tiếng Việt tại ô A20", "vietnamese")
            
            self.source_file_path.set(excel_path)
            
            messagebox.showinfo("Thành công", 
                               f"Đã tạo folder Test và file Excel mẫu!\n"
                               f"Đường dẫn: {test_folder}\n"
                               f"File: Test.xlsx\n\n"
                               f"File đã có dữ liệu tiếng Việt để test.")
            
        except Exception as e:
            self.log_message(f"✗ Lỗi khi tạo folder: {str(e)}", "error")
            messagebox.showerror("Lỗi", f"Không thể tạo folder: {str(e)}")
    
    def enter_vietnamese_text(self, content):
        """Nhập tiếng Việt có dấu vào Excel"""
        method = self.vietnamese_method.get()
        delay = self.step_delay.get()
        
        if method == "paste":
            # Phương pháp 1: Copy vào clipboard rồi paste
            self.log_message("📋 Sử dụng phương pháp Paste (Ctrl+V) cho tiếng Việt", "vietnamese")
            
            # Copy vào clipboard
            pyperclip.copy(content)
            time.sleep(0.5)
            
            # Paste vào Excel
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(delay)
            
            self.log_message(f"✓ Đã paste nội dung tiếng Việt: {content}", "vietnamese")
            
        else:
            # Phương pháp 2: Type từng ký tự (có thể lỗi)
            self.log_message("⌨️ Sử dụng phương pháp Type cho tiếng Việt (có thể lỗi)", "warning")
            
            # Thử type từng ký tự
            try:
                # Chuyển sang bảng mã Unicode
                pyautogui.write(content, interval=0.1)
                time.sleep(delay)
                self.log_message(f"✓ Đã type nội dung: {content}", "vietnamese")
            except Exception as e:
                self.log_message(f"✗ Lỗi khi type tiếng Việt: {str(e)}", "error")
                # Fallback: dùng clipboard
                self.log_message("🔄 Chuyển sang phương pháp Paste...", "warning")
                pyperclip.copy(content)
                time.sleep(0.5)
                pyautogui.hotkey('ctrl', 'v')
                time.sleep(delay)
    
    def focus_excel_window(self, filename):
        """Tập trung vào cửa sổ Excel đang mở"""
        try:
            time.sleep(1)  # Chờ thêm
            
            # Tìm cửa sổ Excel
            for window in gw.getAllWindows():
                if filename.lower() in window.title.lower():
                    if window.isMinimized:
                        window.restore()
                    window.activate()
                    time.sleep(0.5)
                    self.log_message(f"✓ Đã tập trung vào cửa sổ: {window.title}", "success")
                    return True
            
            self.log_message(f"⚠ Không tìm thấy cửa sổ Excel cho: {filename}", "warning")
            return False
                
        except Exception as e:
            self.log_message(f"⚠ Không thể tập trung vào Excel: {str(e)}", "warning")
            return False
    
    def simulate_keyboard_actions(self, file_path):
        """Mô phỏng các thao tác bàn phím với tiếng Việt"""
        try:
            delay = self.step_delay.get()
            cell_to_edit = self.excel_cell.get().upper()
            content_to_write = self.cell_content.get()
            step_by_step = self.show_steps.get()
            
            filename = os.path.basename(file_path)
            
            self.log_message("="*60, "step")
            self.log_message("BẮT ĐẦU THAO TÁC BÀN PHÍM VỚI TIẾNG VIỆT", "step")
            self.log_message("="*60, "step")
            
            if step_by_step:
                messagebox.showinfo("Chuẩn bị", 
                                   f"Sắp bắt đầu thao tác tự động:\n\n"
                                   f"File: {filename}\n"
                                   f"Ô Excel: {cell_to_edit}\n"
                                   f"Nội dung: '{content_to_write}'\n"
                                   f"Phương pháp: {'Paste (Ctrl+V)' if self.vietnamese_method.get() == 'paste' else 'Type'}\n\n"
                                   f"Đảm bảo Excel đang mở và active!")
                time.sleep(2)
            
            # Đợi thêm để chắc chắn Excel đã sẵn sàng
            self.log_message(f"⏳ Đang chờ Excel ổn định...", "warning")
            time.sleep(1)
            
            # Tập trung vào cửa sổ Excel
            self.focus_excel_window(filename)
            time.sleep(0.5)
            
            # Bước 1: Nhấn F5 (Go To)
            self.log_message(f"1. Đang nhấn F5 (Go To)...", "step")
            pyautogui.press('f5')
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 1", "Đã nhấn F5 - Mở hộp thoại Go To")
            
            # Bước 2: Nhập ô Excel cần chỉnh sửa
            self.log_message(f"2. Đang nhập ô Excel: {cell_to_edit}...", "step")
            pyautogui.write(cell_to_edit)
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 2", f"Đã nhập ô Excel: {cell_to_edit}")
            
            # Bước 3: Nhấn Enter để di chuyển đến ô
            self.log_message(f"3. Đang nhấn Enter - Di chuyển đến ô {cell_to_edit}...", "step")
            pyautogui.press('enter')
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 3", f"Đã di chuyển đến ô {cell_to_edit}")
            
            # Bước 4: Nhấn F2 để chỉnh sửa
            self.log_message(f"4. Đang nhấn F2 (Edit mode)...", "step")
            pyautogui.press('f2')
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 4", f"Đã vào chế độ chỉnh sửa ô {cell_to_edit}")
            
            # Bước 5: Xóa nội dung cũ và nhập nội dung mới VỚI TIẾNG VIỆT
            self.log_message(f"5. Đang nhập nội dung tiếng Việt...", "vietnamese")
            
            # Xóa nội dung cũ
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.2)
            pyautogui.press('delete')
            time.sleep(0.2)
            
            # Nhập nội dung mới VỚI TIẾNG VIỆT
            self.enter_vietnamese_text(content_to_write)
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 5", f"Đã nhập nội dung tiếng Việt: '{content_to_write}'")
            
            # Bước 6: Nhấn Enter để hoàn tất
            self.log_message(f"6. Đang nhấn Enter - Lưu thay đổi...", "step")
            pyautogui.press('enter')
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 6", f"Đã lưu thay đổi vào ô {cell_to_edit}")
            
            # Bước 7: Nhấn Ctrl+S để lưu file
            self.log_message(f"7. Đang lưu file (Ctrl+S)...", "step")
            pyautogui.hotkey('ctrl', 's')
            time.sleep(delay)
            
            if step_by_step:
                self.show_step_dialog("Bước 7", "Đã lưu file Excel")
            
            self.log_message("="*60, "step")
            self.log_message(f"✅ HOÀN TẤT! Đã nhập tiếng Việt vào ô {cell_to_edit}", "success")
            self.log_message("="*60, "step")
            
        except Exception as e:
            self.log_message(f"✗ Lỗi khi thao tác bàn phím: {str(e)}", "error")
            raise
            
    def show_step_dialog(self, step_title, step_description):
        """Hiển thị dialog cho từng bước"""
        if self.show_steps.get():
            dialog = tk.Toplevel(self.root)
            dialog.title(f"Bước: {step_title}")
            dialog.geometry("400x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            tk.Label(dialog, text=f"BƯỚC: {step_title}", 
                    font=("Arial", 14, "bold"), fg="blue").pack(pady=10)
            
            tk.Label(dialog, text=step_description, 
                    font=("Arial", 11), wraplength=350).pack(pady=5)
            
            tk.Label(dialog, text="Tiếp tục sau:", 
                    font=("Arial", 10), fg="gray").pack(pady=10)
            
            countdown_label = tk.Label(dialog, text="3", 
                                      font=("Arial", 20, "bold"), fg="red")
            countdown_label.pack()
            
            def update_countdown(count):
                if count > 0:
                    countdown_label.config(text=str(count))
                    dialog.after(1000, update_countdown, count-1)
                else:
                    dialog.destroy()
            
            dialog.after(100, update_countdown, 3)
            dialog.wait_window()
    
    def test_vietnamese_input(self):
        """Test nhập tiếng Việt"""
        try:
            content = self.cell_content.get()
            self.log_message(f"🧪 Bắt đầu test nhập tiếng Việt: '{content}'", "vietnamese")
            
            # Copy vào clipboard để test
            pyperclip.copy(content)
            time.sleep(0.5)
            
            # Mở notepad để test
            messagebox.showinfo("Test tiếng Việt", 
                              f"Đã copy vào clipboard: '{content}'\n\n"
                              f"Mở Notepad và nhấn Ctrl+V để test.\n"
                              f"Hoặc mở Excel và test paste.")
            
            self.log_message(f"✓ Đã copy vào clipboard: {content}", "vietnamese")
            self.log_message("📝 Mở Notepad/Excel và nhấn Ctrl+V để test", "warning")
            
        except Exception as e:
            self.log_message(f"✗ Lỗi khi test tiếng Việt: {str(e)}", "error")
            
    def test_keyboard_actions(self):
        """Test thao tác bàn phím"""
        try:
            self.log_message("🎯 Bắt đầu test thao tác bàn phím...", "warning")
            
            confirm = messagebox.askyesno("Test thao tác", 
                                         "Bạn muốn test thao tác bàn phím?\n\n"
                                         "Đảm bảo Excel đang mở và active!\n"
                                         f"Ô Excel: {self.excel_cell.get()}\n"
                                         f"Nội dung: '{self.cell_content.get()}'\n"
                                         f"Phương pháp: {'Paste' if self.vietnamese_method.get() == 'paste' else 'Type'}")
            
            if confirm:
                thread = threading.Thread(
                    target=lambda: self.simulate_keyboard_actions("Test.xlsx"), 
                    daemon=True
                )
                thread.start()
                
        except Exception as e:
            self.log_message(f"✗ Lỗi khi test: {str(e)}", "error")
            
    def copy_file_only(self):
        """Chỉ copy file mà không thực hiện RPA"""
        try:
            self.log_message("📋 Bắt đầu copy file...", "step")
            new_file_path = self.copy_and_rename_file()
            
            if new_file_path:
                self.new_file_path = new_file_path
                self.log_message(f"✅ Đã copy file thành công: {os.path.basename(new_file_path)}", "success")
                messagebox.showinfo("Thành công", 
                                  f"Đã copy file thành công!\n\n"
                                  f"File: {os.path.basename(new_file_path)}\n"
                                  f"Đường dẫn: {os.path.dirname(new_file_path)}")
                
        except Exception as e:
            self.log_message(f"✗ Lỗi khi copy file: {str(e)}", "error")
            messagebox.showerror("Lỗi", f"Không thể copy file: {str(e)}")
            
    def copy_and_rename_file(self):
        """Copy và đổi tên file"""
        try:
            source_path = self.source_file_path.get()
            dest_folder = self.destination_folder.get()
            
            if not source_path or not os.path.exists(source_path):
                messagebox.showerror("Lỗi", "Vui lòng chọn file nguồn hợp lệ!")
                return None
                
            if not dest_folder:
                messagebox.showerror("Lỗi", "Vui lòng chọn thư mục đích!")
                return None
            
            current_date = datetime.now().strftime("%d-%m-%Y")
            new_folder_name = f"Test_{current_date}"
            new_folder_path = os.path.join(dest_folder, new_folder_name)
            
            os.makedirs(new_folder_path, exist_ok=True)
            self.log_message(f"📁 Đã tạo folder: {new_folder_name}", "success")
            
            new_filename = f"Test_{current_date}.xlsx"
            new_file_path = os.path.join(new_folder_path, new_filename)
            
            shutil.copy2(source_path, new_file_path)
            self.log_message(f"📄 Đã copy file thành: {new_filename}", "success")
            
            return new_file_path
            
        except Exception as e:
            self.log_message(f"✗ Lỗi khi copy file: {str(e)}", "error")
            raise
            
    def open_excel_file_with_wait(self, file_path):
        """Mở file Excel và chờ đủ thời gian"""
        try:
            if file_path and os.path.exists(file_path):
                filename = os.path.basename(file_path)
                
                self.log_message(f"🔵 Đang mở file Excel: {filename}...", "step")
                
                if os.name == 'nt':
                    os.startfile(file_path)
                else:
                    subprocess.call(['open', file_path])
                
                wait_time = self.open_wait_time.get()
                self.log_message(f"⏳ Đang chờ {wait_time} giây cho Excel mở hoàn toàn...", "warning")
                
                for i in range(int(wait_time), 0, -1):
                    self.log_message(f"   Còn {i} giây...", "warning")
                    time.sleep(1)
                
                self.log_message(f"✅ File Excel đã mở: {filename}", "success")
                
                time.sleep(0.5)
                self.focus_excel_window(filename)
                
                return True
            else:
                self.log_message(f"✗ File không tồn tại: {file_path}", "error")
                return False
                
        except Exception as e:
            self.log_message(f"✗ Lỗi khi mở file Excel: {str(e)}", "error")
            return False
            
    def open_excel_file(self):
        """Mở file Excel (cho nút riêng)"""
        file_path = self.source_file_path.get()
        if file_path and os.path.exists(file_path):
            try:
                self.open_excel_file_with_wait(file_path)
            except Exception as e:
                self.log_message(f"✗ Lỗi khi mở file: {str(e)}", "error")
        else:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel hợp lệ trước!")
            
    def run_rpa_in_thread(self):
        """Chạy RPA trong thread riêng với đúng thứ tự"""
        try:
            self.start_button.config(state="disabled")
            
            self.log_message("🚀 BẮT ĐẦU QUÁ TRÌNH RPA ĐẦY ĐỦ...", "step")
            self.log_message("="*60, "step")
            
            # Bước 1: Copy và đổi tên file
            self.log_message("📋 Bước 1: Copy và đổi tên file...", "step")
            new_file_path = self.copy_and_rename_file()
            
            if not new_file_path:
                return
            
            self.new_file_path = new_file_path
            
            # Bước 2: Mở file Excel mới và CHỜ ĐỦ THỜI GIAN
            self.log_message("📂 Bước 2: Mở file Excel mới...", "step")
            excel_opened = self.open_excel_file_with_wait(new_file_path)
            
            if not excel_opened:
                self.log_message("✗ Không thể mở file Excel, dừng RPA", "error")
                return
            
            # Bước 3: Mô phỏng thao tác bàn phím với TIẾNG VIỆT
            self.log_message("⌨️ Bước 3: Thực hiện thao tác bàn phím với TIẾNG VIỆT...", "vietnamese")
            self.simulate_keyboard_actions(new_file_path)
            
            self.log_message("="*60, "step")
            self.log_message("✅ QUÁ TRÌNH RPA HOÀN TẤT THÀNH CÔNG!", "success")
            self.log_message("="*60, "step")
            
            messagebox.showinfo("Thành công", 
                              f"RPA process completed successfully!\n\n"
                              f"📁 Folder mới: Test_{datetime.now().strftime('%d-%m-%Y')}\n"
                              f"📄 File mới: Test_{datetime.now().strftime('%d-%m-%Y')}.xlsx\n"
                              f"📍 Ô Excel: {self.excel_cell.get()}\n"
                              f"📝 Nội dung: '{self.cell_content.get()}'\n"
                              f"🔤 Phương pháp: {'Paste (Ctrl+V)' if self.vietnamese_method.get() == 'paste' else 'Type'}\n\n"
                              f"File đã được lưu với thay đổi.")
            
        except Exception as e:
            self.log_message(f"❌ Lỗi trong quá trình RPA: {str(e)}", "error")
            messagebox.showerror("Lỗi", f"RPA process failed: {str(e)}")
            
        finally:
            self.start_button.config(state="normal")
            
    def start_rpa_process(self):
        """Bắt đầu quá trình RPA"""
        if not self.source_file_path.get():
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file nguồn trước!")
            return
            
        if not self.destination_folder.get():
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn thư mục đích trước!")
            return
        
        if not self.cell_content.get().strip():
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập nội dung cần ghi vào Excel!")
            return
            
        if not self.excel_cell.get().strip():
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập ô Excel (ví dụ: A20)!")
            return
        
        confirm_msg = (
            f"XÁC NHẬN CHẠY RPA - VỚI TIẾNG VIỆT\n\n"
            f"1. COPY FILE:\n"
            f"   • File nguồn: {os.path.basename(self.source_file_path.get())}\n"
            f"   • Folder mới: Test_{datetime.now().strftime('%d-%m-%Y')}\n\n"
            f"2. MỞ FILE EXCEL VÀ CHỜ:\n"
            f"   • Thời gian chờ: {self.open_wait_time.get()} giây\n\n"
            f"3. THAO TÁC BÀN PHÍM VỚI TIẾNG VIỆT:\n"
            f"   • Ô Excel: {self.excel_cell.get()}\n"
            f"   • Nội dung: '{self.cell_content.get()}'\n"
            f"   • Phương pháp: {'PASTE (Ctrl+V)' if self.vietnamese_method.get() == 'paste' else 'TYPE'}\n"
            f"   • Hiển thị từng bước: {'CÓ' if self.show_steps.get() else 'KHÔNG'}\n\n"
            f"⚠ LƯU Ý: Sau khi Excel mở, KHÔNG sử dụng chuột/bàn phím!"
        )
        
        confirm = messagebox.askyesno("Xác nhận chạy RPA - TIẾNG VIỆT", confirm_msg)
        
        if confirm:
            thread = threading.Thread(target=self.run_rpa_in_thread, daemon=True)
            thread.start()

def main():
    root = tk.Tk()
    app = ExcelRPACopier(root)
    root.mainloop()

if __name__ == "__main__":
    main()